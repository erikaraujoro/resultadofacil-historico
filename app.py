import os
import re
import time
import json
import logging
import threading
from pathlib import Path
from datetime import datetime, timedelta
from io import BytesIO

import requests
from bs4 import BeautifulSoup
from flask import Flask, jsonify, send_file
from openpyxl import Workbook


app = Flask(__name__)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)


# ==========================================================
# CONFIGURAÇÕES
# ==========================================================

LOTERIAS = {
    "PT-SP": {
        "slug": "resultados-pt-sp-do-dia-",
        "url_base": "https://www.resultadofacil.com.br/",
    },

    "LOTEP": {
        "slug": "resultados-lotep-do-dia-",
        "url_base": "https://www.resultadofacil.com.br/",
    },

    "PT-BA": {
        "slug": "resultados-paratodos-bahia-do-dia-",
        "url_base": "https://www.resultadofacil.com.br/",
    },

    "LOTECE": {
        "slug": "resultados-lotece---loteria-dos-sonhos-do-dia-",
        "url_base": "https://www.resultadofacil.com.br/",
    },
}


HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/142.0 Safari/537.36"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
}

# ==========================================================
# JB CERTO - DIAGNÓSTICO HISTÓRICO LOTEP 20H
# ==========================================================

URL_JBCERTO_LOTEP = (
    "https://resultadosjbcerto.com.br/lotep/"
)

# ==========================================================
# COLETA HISTÓRICA / DISCO PERSISTENTE
# ==========================================================

DATA_INICIAL_COLETA = datetime(
    2025,
    1,
    1
)

DATA_FINAL_COLETA = datetime(
    2026,
    8,
    12
)

DIRETORIO_DADOS = Path(
    os.environ.get(
        "DATA_DIR",
        "/var/data"
    )
)

DIRETORIO_DADOS.mkdir(
    parents=True,
    exist_ok=True
)

ARQUIVO_ESTADO = (
    DIRETORIO_DADOS
    / "estado_coleta.json"
)

ARQUIVO_RESULTADOS = (
    DIRETORIO_DADOS
    / "resultados.jsonl"
)

ARQUIVO_AUDITORIA = (
    DIRETORIO_DADOS
    / "auditoria.jsonl"
)

ARQUIVO_EXCEL = (
    DIRETORIO_DADOS
    / "resultadofacil_2025_2026.xlsx"
)


# ==========================================================
# COLETA COMPLEMENTAR LOTEP 20H
# Fonte: Resultado Fácil - bloco PARATODOS/PB 20h
# Período: 01/01/2025 a 12/08/2026
# ==========================================================

LOTEP20_DATA_INICIAL = datetime(2025, 1, 1)
LOTEP20_DATA_FINAL = datetime(2026, 8, 12)

LOTEP20_URL_BASE = "https://www.resultadofacil.com.br/"

ARQUIVO_LOTEP20_ESTADO = DIRETORIO_DADOS / "lotep20_estado.json"
ARQUIVO_LOTEP20_RESULTADOS = DIRETORIO_DADOS / "lotep20_resultados.jsonl"
ARQUIVO_LOTEP20_AUDITORIA = DIRETORIO_DADOS / "lotep20_auditoria.jsonl"
ARQUIVO_LOTEP20_EXCEL = DIRETORIO_DADOS / "lotep_09_20_2025_2026.xlsx"

LOTEP20_LOCK = threading.Lock()
LOTEP20_THREAD = None

LOTEP20_TOTAL_DIAS = (
    LOTEP20_DATA_FINAL.date()
    - LOTEP20_DATA_INICIAL.date()
).days + 1


# ==========================================================
# COLETA COMPLEMENTAR PT-SP 15H
# Fonte: Resultado Fácil - banca BANDEIRANTES
# Período: 01/01/2025 a 11/08/2026
# ==========================================================

PTSP15_DATA_INICIAL = datetime(2025, 1, 1)
PTSP15_DATA_FINAL = datetime(2026, 8, 11)
PTSP15_URL_BASE = "https://www.resultadofacil.com.br/"

ARQUIVO_PTSP15_ESTADO = DIRETORIO_DADOS / "ptsp15_estado.json"
ARQUIVO_PTSP15_RESULTADOS = DIRETORIO_DADOS / "ptsp15_resultados.jsonl"
ARQUIVO_PTSP15_AUDITORIA = DIRETORIO_DADOS / "ptsp15_auditoria.jsonl"
ARQUIVO_PTSP15_EXCEL = DIRETORIO_DADOS / "ptsp_15h_2025_2026.xlsx"

PTSP15_LOCK = threading.Lock()
PTSP15_THREAD = None

PTSP15_TOTAL_DIAS = (
    PTSP15_DATA_FINAL.date()
    - PTSP15_DATA_INICIAL.date()
).days + 1


# ==========================================================
# NOVA COLETA HISTÓRICA
# AVAL-PE + CAMINHO-DA-SORTE + MINAS-MG
# Período: 01/01/2025 a 14/08/2026
#
# Regras especiais:
# AVAL-PE: origem 13 -> grava 12 | origem 16 -> grava 15
# CAMINHO-DA-SORTE: origem 20 -> grava 19
# MINAS-MG: aceita ALVORADA 12, MINAS DIA 15,
#           MINAS NOITE 19 e PREFERIDA 21
# Federal é sempre ignorada.
# ==========================================================

NOVAS_LOTERIAS = {
    "AVAL-PE": {
        "slug": "resultados-aval-pernambuco-do-dia-",
        "url_base": "https://www.resultadofacil.com.br/",
        # No Dataset do Resultado Fácil:
        # 12:45 aparece com horário efetivo 13:00
        # 15:45 aparece com horário efetivo 16:00
        "horarios_origem": {"09", "11", "13", "14", "16", "17", "19"},
        "normalizacao_horarios": {
            "13": "12",
            "16": "15",
        },
    },
    "CAMINHO-DA-SORTE": {
        "slug": "resultados-caminho-da-sorte-do-dia-",
        "url_base": "https://www.resultadofacil.com.br/",
        "horarios_origem": {"09", "11", "12", "14", "15", "17", "18", "20", "21"},
        "normalizacao_horarios": {"20": "19"},
    },
    "MINAS-MG": {
        "slug": "resultados-minas-mg-do-dia-",
        "url_base": "https://www.resultadofacil.com.br/",
        "horarios_origem": {"12", "15", "19", "21"},
        "normalizacao_horarios": {},
    },
}

NOVAS_DATA_INICIAL = datetime(2025, 1, 1)
NOVAS_DATA_FINAL = datetime(2026, 8, 14)

ARQUIVO_NOVAS_ESTADO = DIRETORIO_DADOS / "novas_loterias_estado.json"
ARQUIVO_NOVAS_RESULTADOS = DIRETORIO_DADOS / "novas_loterias_resultados.jsonl"
ARQUIVO_NOVAS_AUDITORIA = DIRETORIO_DADOS / "novas_loterias_auditoria.jsonl"
ARQUIVO_NOVAS_EXCEL = DIRETORIO_DADOS / "aval_pe_caminho_da_sorte_minas_mg_2025_2026.xlsx"

NOVAS_LOCK = threading.Lock()
NOVAS_THREAD = None

NOVAS_TOTAL_DIAS = (
    NOVAS_DATA_FINAL.date()
    - NOVAS_DATA_INICIAL.date()
).days + 1

NOVAS_TOTAL_PAGINAS = (
    NOVAS_TOTAL_DIAS
    * len(NOVAS_LOTERIAS)
)

COLETA_LOCK = threading.Lock()

COLETA_THREAD = None

TOTAL_DIAS_COLETA = (
    DATA_FINAL_COLETA.date()
    - DATA_INICIAL_COLETA.date()
).days + 1

TOTAL_PAGINAS_COLETA = (
    TOTAL_DIAS_COLETA
    * len(LOTERIAS)
)

def estado_inicial_coleta():
    return {
        "status": "nao_iniciada",
        "data_inicial": DATA_INICIAL_COLETA.strftime(
            "%d/%m/%Y"
        ),
        "data_final": DATA_FINAL_COLETA.strftime(
            "%d/%m/%Y"
        ),
        "data_atual": None,
        "loteria_atual": None,
        "paginas_processadas": 0,
        "resultados_coletados": 0,
        "falhas": 0,
        "progresso": 0.0,
        "mensagem": (
            "A coleta ainda não foi iniciada."
        ),
    }


def salvar_estado_coleta(estado):
    temporario = ARQUIVO_ESTADO.with_suffix(
        ".tmp"
    )

    with open(
        temporario,
        "w",
        encoding="utf-8"
    ) as arquivo:
        json.dump(
            estado,
            arquivo,
            ensure_ascii=False,
            indent=2,
        )

    os.replace(
        temporario,
        ARQUIVO_ESTADO
    )


def carregar_estado_coleta():
    if not ARQUIVO_ESTADO.exists():
        estado = estado_inicial_coleta()
        salvar_estado_coleta(
            estado
        )
        return estado

    try:
        with open(
            ARQUIVO_ESTADO,
            "r",
            encoding="utf-8"
        ) as arquivo:
            return json.load(
                arquivo
            )

    except Exception:
        logging.exception(
            "Erro ao carregar estado da coleta."
        )

        return estado_inicial_coleta()


def adicionar_jsonl(
    caminho,
    registro
):
    with open(
        caminho,
        "a",
        encoding="utf-8"
    ) as arquivo:
        arquivo.write(
            json.dumps(
                registro,
                ensure_ascii=False,
            )
        )
        arquivo.write("\n")
        
def carregar_chaves_resultados():
    chaves = set()

    if not ARQUIVO_RESULTADOS.exists():
        return chaves

    try:
        with open(
            ARQUIVO_RESULTADOS,
            "r",
            encoding="utf-8"
        ) as arquivo:

            for linha in arquivo:
                linha = linha.strip()

                if not linha:
                    continue

                try:
                    registro = json.loads(
                        linha
                    )
                except Exception:
                    continue

                chave = (
                    f"{registro.get('data', '')}|"
                    f"{registro.get('loteria', '')}|"
                    f"{registro.get('horario', '')}"
                )

                chaves.add(
                    chave
                )

    except Exception:
        logging.exception(
            "Erro ao carregar chaves dos resultados."
        )

    return chaves


def carregar_paginas_concluidas():
    """
    Retorna as páginas já processadas com sucesso.

    Uma página é identificada por:
    DATA + LOTERIA

    Falhas HTTP/rede não entram aqui,
    pois deverão ser tentadas novamente.
    """

    concluidas = set()

    if not ARQUIVO_AUDITORIA.exists():
        return concluidas

    try:
        with open(
            ARQUIVO_AUDITORIA,
            "r",
            encoding="utf-8"
        ) as arquivo:

            for linha in arquivo:
                linha = linha.strip()

                if not linha:
                    continue

                try:
                    registro = json.loads(
                        linha
                    )
                except Exception:
                    continue

                status = registro.get(
                    "status",
                    ""
                )

                if status not in {
                    "ok",
                    "nao_encontrado",
                    "sem_resultados_validos",
                }:
                    continue

                chave = (
                    f"{registro.get('data', '')}|"
                    f"{registro.get('loteria', '')}"
                )

                concluidas.add(
                    chave
                )

    except Exception:
        logging.exception(
            "Erro ao carregar auditoria."
        )

    return concluidas

def executar_coleta_historica():
    global COLETA_THREAD

    logging.info(
        "Iniciando coleta histórica Resultado Fácil."
    )

    chaves_resultados = (
        carregar_chaves_resultados()
    )

    paginas_concluidas = (
        carregar_paginas_concluidas()
    )

    estado = carregar_estado_coleta()

    estado.update({
        "status": "executando",
        "data_atual": None,
        "loteria_atual": None,
        "paginas_processadas": len(
            paginas_concluidas
        ),
        "resultados_coletados": len(
            chaves_resultados
        ),
        "falhas": 0,
        "progresso": round(
            (
                len(paginas_concluidas)
                / TOTAL_PAGINAS_COLETA
            ) * 100,
            2
        ),
        "mensagem": (
            "Coleta histórica em andamento."
        ),
    })

    salvar_estado_coleta(
        estado
    )

    data_atual = (
        DATA_INICIAL_COLETA
    )

    try:
        while (
            data_atual
            <= DATA_FINAL_COLETA
        ):

            data_br = data_atual.strftime(
                "%d/%m/%Y"
            )

            for loteria in LOTERIAS.keys():

                chave_pagina = (
                    f"{data_br}|{loteria}"
                )

                # Página já processada anteriormente
                if (
                    chave_pagina
                    in paginas_concluidas
                ):
                    continue

                estado[
                    "data_atual"
                ] = data_br

                estado[
                    "loteria_atual"
                ] = loteria

                estado[
                    "mensagem"
                ] = (
                    f"Consultando "
                    f"{loteria} em "
                    f"{data_br}."
                )

                salvar_estado_coleta(
                    estado
                )

                logging.info(
                    "Consultando %s | %s",
                    data_br,
                    loteria,
                )

                try:
                    retorno = (
                        buscar_resultados_data(
                            loteria,
                            data_atual
                        )
                    )

                    status = retorno.get(
                        "status",
                        "desconhecido"
                    )

                    resultados = retorno.get(
                        "resultados",
                        []
                    )

                    quantidade_novos = 0

                    if status == "ok":

                        for resultado in resultados:

                            chave_resultado = (
                                f"{resultado['data']}|"
                                f"{resultado['loteria']}|"
                                f"{resultado['horario']}"
                            )

                            if (
                                chave_resultado
                                in chaves_resultados
                            ):
                                continue

                            adicionar_jsonl(
                                ARQUIVO_RESULTADOS,
                                resultado
                            )

                            chaves_resultados.add(
                                chave_resultado
                            )

                            quantidade_novos += 1

                    registro_auditoria = {
                        "data": data_br,
                        "loteria": loteria,
                        "status": status,
                        "quantidade": len(
                            resultados
                        ),
                        "novos": quantidade_novos,
                        "url": retorno.get(
                            "url",
                            ""
                        ),
                        "erro": retorno.get(
                            "erro",
                            ""
                        ),
                        "processado_em": datetime.now().strftime(
                            "%d/%m/%Y %H:%M:%S"
                        ),
                    }

                    adicionar_jsonl(
                        ARQUIVO_AUDITORIA,
                        registro_auditoria
                    )

                    # Só marca como concluída
                    # se não houve erro de rede/HTTP.
                    if status in {
                        "ok",
                        "nao_encontrado",
                        "sem_resultados_validos",
                    }:
                        paginas_concluidas.add(
                            chave_pagina
                        )

                    else:
                        estado["falhas"] += 1

                except Exception as e:

                    logging.exception(
                        "Erro durante coleta %s | %s",
                        data_br,
                        loteria,
                    )

                    estado[
                        "falhas"
                    ] += 1

                    adicionar_jsonl(
                        ARQUIVO_AUDITORIA,
                        {
                            "data": data_br,
                            "loteria": loteria,
                            "status": "erro_interno",
                            "quantidade": 0,
                            "novos": 0,
                            "url": montar_url(
                                loteria,
                                data_atual
                            ),
                            "erro": str(e),
                            "processado_em": datetime.now().strftime(
                                "%d/%m/%Y %H:%M:%S"
                            ),
                        }
                    )

                estado[
                    "paginas_processadas"
                ] = len(
                    paginas_concluidas
                )

                estado[
                    "resultados_coletados"
                ] = len(
                    chaves_resultados
                )

                estado[
                    "progresso"
                ] = round(
                    (
                        len(
                            paginas_concluidas
                        )
                        / TOTAL_PAGINAS_COLETA
                    ) * 100,
                    2
                )

                salvar_estado_coleta(
                    estado
                )

                # Pequena pausa para não bombardear
                # o Resultado Fácil.
                time.sleep(0.8)

            data_atual += timedelta(
                days=1
            )

        estado.update({
            "status": "concluida",
            "data_atual": DATA_FINAL_COLETA.strftime(
                "%d/%m/%Y"
            ),
            "loteria_atual": None,
            "paginas_processadas": len(
                paginas_concluidas
            ),
            "resultados_coletados": len(
                chaves_resultados
            ),
            "progresso": round(
                (
                    len(
                        paginas_concluidas
                    )
                    / TOTAL_PAGINAS_COLETA
                ) * 100,
                2
            ),
            "mensagem": (
                "Coleta histórica concluída."
            ),
        })

        salvar_estado_coleta(
            estado
        )

        logging.info(
            "Coleta histórica concluída. "
            "Resultados: %s | "
            "Páginas: %s | "
            "Falhas: %s",
            len(chaves_resultados),
            len(paginas_concluidas),
            estado["falhas"],
        )

    except Exception as e:

        logging.exception(
            "Falha geral na coleta histórica."
        )

        estado.update({
            "status": "erro",
            "mensagem": str(e),
        })

        salvar_estado_coleta(
            estado
        )

    finally:
        with COLETA_LOCK:
            COLETA_THREAD = None


# ==========================================================
# UTILITÁRIOS
# ==========================================================

def somente_digitos(valor):
    return re.sub(r"\D", "", str(valor or ""))


def formatar_milhar(valor):
    digitos = somente_digitos(valor)

    if not digitos:
        return ""

    return digitos[-4:].zfill(4)


def normalizar_texto(texto):
    return re.sub(
        r"\s+",
        " ",
        str(texto or "")
    ).strip()


def calcular_premios_6_7(premios):
    valores = [int(p) for p in premios[:5]]

    soma = sum(valores)

    m6 = str(soma)[-4:].zfill(4)

    produto = valores[0] * valores[1]

    m7 = str(
        produto // 1000
    )[-3:].zfill(3)

    return m6, m7


def montar_url(loteria, data_obj):
    cfg = LOTERIAS[loteria]

    data_iso = data_obj.strftime(
        "%Y-%m-%d"
    )

    return (
        cfg["url_base"]
        + cfg["slug"]
        + data_iso
    )


# ==========================================================
# EXTRAÇÃO
# ==========================================================

def extrair_horario(texto):
    texto = normalizar_texto(
        texto
    ).lower()

    match = re.search(
        r"\b(\d{1,2})\s*(?:h|horas?)",
        texto
    )

    if match:
        return match.group(1).zfill(2)

    match = re.search(
        r"\b(\d{1,2}):\d{2}\b",
        texto
    )

    if match:
        return match.group(1).zfill(2)

    return ""


def parece_federal(texto):
    texto = normalizar_texto(
        texto
    ).lower()

    return "federal" in texto

def obter_campo_item(item, nomes):
    """
    O Resultado Fácil apresenta pequenas variações no JSON-LD.
    Alguns registros usam name/value e outros podem aparecer
    como nome/valor.

    Esta função aceita essas variações.
    """

    for nome in nomes:
        valor = item.get(nome)

        if valor not in [None, ""]:
            return str(valor).strip()

    return ""


def extrair_dataset_resultadofacil(html):
    """
    Localiza o Dataset de resultados dentro dos blocos JSON-LD.
    """

    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    scripts = soup.find_all(
        "script",
        attrs={"type": "application/ld+json"}
    )

    for script in scripts:
        conteudo = script.string

        if not conteudo:
            conteudo = script.get_text(
                strip=True
            )

        if not conteudo:
            continue

        try:
            dados = json.loads(
                conteudo
            )
        except Exception:
            continue

        grafo = dados.get(
            "@graph",
            []
        )

        if not isinstance(
            grafo,
            list
        ):
            continue

        for item in grafo:

            if not isinstance(
                item,
                dict
            ):
                continue

            tipo = str(
                item.get(
                    "@type",
                    ""
                )
            ).lower()

            if tipo != "dataset":
                continue

            variaveis = item.get(
                "variableMeasured",
                []
            )

            if isinstance(
                variaveis,
                list
            ):
                return {
                    "dataset": item,
                    "variaveis": variaveis,
                }

    return {
        "dataset": None,
        "variaveis": [],
    }


def normalizar_nome_resultado(texto):
    texto = normalizar_texto(
        texto
    )

    texto = texto.replace(
        "º",
        "o"
    )

    texto = texto.replace(
        "ª",
        "a"
    )

    return texto


def extrair_posicao_premio(texto):
    """
    Identifica somente posições de 1º a 5º prêmio.
    """

    texto = normalizar_nome_resultado(
        texto
    ).lower()

    match = re.search(
        r"\b([1-5])\s*[oa]?\s*pr[eê]mio\b",
        texto
    )

    if not match:
        return None

    return int(
        match.group(1)
    )


def extrair_horario_variavel(texto):
    """
    Obtém o horário do nome da variável do Dataset.

    Exemplos:
    PTSP 08:20
    BA 10:20
    LOTEP 10:45
    CE 15:45
    """

    texto = normalizar_texto(
        texto
    )

    horarios = re.findall(
        r"\b([01]?\d|2[0-3]):([0-5]\d)\b",
        texto
    )

    if horarios:
        hora = horarios[-1][0]

        return hora.zfill(2)

    match = re.search(
        r"\b(\d{1,2})\s*h\b",
        texto.lower()
    )

    if match:
        return match.group(1).zfill(2)

    return ""


def pertence_loteria(
    loteria,
    nome_variavel
):
    """
    Impede que resultados da Federal ou de outra banca
    sejam incorporados à loteria analisada.
    """

    texto = normalizar_texto(
        nome_variavel
    ).upper()

    # Federal misturada nas páginas
    if "FEDERAL" in texto:
        return False

    if loteria == "PT-SP":
        return (
            "PTSP" in texto
            or "PT SP" in texto
            or "PTN SP" in texto
        )

    if loteria == "LOTEP":
        return "LOTEP" in texto

    if loteria == "PT-BA":
        return (
            " BA " in f" {texto} "
            or texto.startswith("BA ")
            or " - BA" in texto
        )

    if loteria == "LOTECE":
        return (
            "LOTECE" in texto
            or " CE," in texto
            or " CE " in texto
        )

    return False


def extrair_milhar_valor(valor):
    """
    O campo value normalmente vem assim:

    5388 · Grupo 22 · Tigre

    Captura apenas a primeira milhar.
    """

    valor = normalizar_texto(
        valor
    )

    match = re.search(
        r"(?<!\d)(\d{1,4})(?!\d)",
        valor
    )

    if not match:
        return ""

    return match.group(1).zfill(4)


def extrair_resultados_dataset(
    html,
    loteria,
    data_obj,
    url
):
    estrutura = extrair_dataset_resultadofacil(
        html
    )

    variaveis = estrutura[
        "variaveis"
    ]

    sorteios = {}

    for item in variaveis:

        if not isinstance(
            item,
            dict
        ):
            continue

        nome = obter_campo_item(
            item,
            [
                "name",
                "nome",
            ]
        )

        valor = obter_campo_item(
            item,
            [
                "value",
                "valor",
            ]
        )

        if not nome or not valor:
            continue

        if not pertence_loteria(
            loteria,
            nome
        ):
            continue

        posicao = extrair_posicao_premio(
            nome
        )

        if posicao is None:
            continue

        horario = extrair_horario_variavel(
            nome
        )

        if not horario:
            continue

        milhar = extrair_milhar_valor(
            valor
        )

        if not milhar:
            continue

        if horario not in sorteios:
            sorteios[horario] = {}

        sorteios[horario][
            posicao
        ] = milhar

    resultados = []

    for horario in sorted(
        sorteios.keys()
    ):

        premios_dict = sorteios[
            horario
        ]

        # Só aceita sorteio completo do 1º ao 5º
        if not all(
            p in premios_dict
            for p in range(1, 6)
        ):
            continue

        premios = [
            premios_dict[1],
            premios_dict[2],
            premios_dict[3],
            premios_dict[4],
            premios_dict[5],
        ]

        m6, m7 = calcular_premios_6_7(
            premios
        )

        resultados.append({
            "data": data_obj.strftime(
                "%d/%m/%Y"
            ),
            "loteria": loteria,
            "horario": horario,
            "m1": premios[0],
            "m2": premios[1],
            "m3": premios[2],
            "m4": premios[3],
            "m5": premios[4],
            "m6": m6,
            "m7": m7,
            "url": url,
        })

    return resultados

def buscar_resultados_data(
    loteria,
    data_obj
):
    url = montar_url(
        loteria,
        data_obj
    )

    try:
        resp = requests.get(
            url,
            headers=HEADERS,
            timeout=30,
        )

    except Exception as e:
        return {
            "ok": False,
            "status": "erro_rede",
            "erro": str(e),
            "url": url,
            "resultados": [],
        }

    if resp.status_code == 404:
        return {
            "ok": True,
            "status": "nao_encontrado",
            "url": url,
            "resultados": [],
        }

    if resp.status_code == 403:
        return {
            "ok": False,
            "status": "bloqueado_403",
            "url": url,
            "resultados": [],
        }

    if resp.status_code != 200:
        return {
            "ok": False,
            "status": f"http_{resp.status_code}",
            "url": url,
            "resultados": [],
        }

    resultados = extrair_resultados_dataset(
        resp.text,
        loteria,
        data_obj,
        url,
    )

    if not resultados:
        return {
            "ok": True,
            "status": "sem_resultados_validos",
            "url": url,
            "resultados": [],
        }

    return {
        "ok": True,
        "status": "ok",
        "url": url,
        "resultados": resultados,
    }

def diagnosticar_pagina_resultadofacil(loteria, data_obj):
    url = montar_url(
        loteria,
        data_obj
    )

    resp = requests.get(
        url,
        headers=HEADERS,
        timeout=30,
    )

    return {
        "status_http": resp.status_code,
        "url": url,
        "html": resp.text[:50000],
    }
    

def carregar_jsonl(caminho):
    registros = []

    if not caminho.exists():
        return registros

    with open(
        caminho,
        "r",
        encoding="utf-8"
    ) as arquivo:

        for linha in arquivo:
            linha = linha.strip()

            if not linha:
                continue

            try:
                registro = json.loads(
                    linha
                )

                registros.append(
                    registro
                )

            except Exception:
                logging.exception(
                    "Linha inválida em %s",
                    caminho
                )

    return registros

def gerar_excel_historico():
    resultados = carregar_jsonl(
        ARQUIVO_RESULTADOS
    )

    auditoria = carregar_jsonl(
        ARQUIVO_AUDITORIA
    )

    if not resultados:
        raise ValueError(
            "Nenhum resultado coletado foi encontrado."
        )

    # ======================================================
    # ORDENAÇÃO DOS RESULTADOS
    # ======================================================

    ordem_loterias = {
        "PT-SP": 1,
        "LOTEP": 2,
        "PT-BA": 3,
        "LOTECE": 4,
    }

    def chave_resultado(item):
        try:
            data = datetime.strptime(
                item.get("data", ""),
                "%d/%m/%Y"
            )
        except Exception:
            data = datetime.max

        horario = str(
            item.get(
                "horario",
                ""
            )
        ).zfill(2)

        return (
            data,
            ordem_loterias.get(
                item.get("loteria", ""),
                999
            ),
            horario,
        )

    resultados.sort(
        key=chave_resultado
    )

    # ======================================================
    # ORDENAÇÃO DA AUDITORIA
    # ======================================================

    def chave_auditoria(item):
        try:
            data = datetime.strptime(
                item.get("data", ""),
                "%d/%m/%Y"
            )
        except Exception:
            data = datetime.max

        return (
            data,
            ordem_loterias.get(
                item.get("loteria", ""),
                999
            ),
        )

    auditoria.sort(
        key=chave_auditoria
    )

    # ======================================================
    # CRIA WORKBOOK
    # ======================================================

    wb = Workbook()

    ws = wb.active
    ws.title = "RESULTADOS"

    cabecalho_resultados = [
        "Data",
        "Loteria",
        "Horário",
        "M1",
        "M2",
        "M3",
        "M4",
        "M5",
        "M6",
        "M7",
    ]

    ws.append(
        cabecalho_resultados
    )

    for resultado in resultados:

        ws.append([
            resultado.get("data", ""),
            resultado.get("loteria", ""),
            resultado.get("horario", ""),
            resultado.get("m1", ""),
            resultado.get("m2", ""),
            resultado.get("m3", ""),
            resultado.get("m4", ""),
            resultado.get("m5", ""),
            resultado.get("m6", ""),
            resultado.get("m7", ""),
        ])

    # ======================================================
    # PRESERVA ZEROS À ESQUERDA
    # ======================================================

    for linha in range(
        2,
        ws.max_row + 1
    ):

        # Horário
        ws.cell(
            linha,
            3
        ).number_format = "@"

        # M1 até M6 = 4 caracteres
        for coluna in range(
            4,
            10
        ):
            celula = ws.cell(
                linha,
                coluna
            )

            celula.value = str(
                celula.value
            ).zfill(4)

            celula.number_format = "@"

        # M7 = 3 caracteres
        celula_m7 = ws.cell(
            linha,
            10
        )

        celula_m7.value = str(
            celula_m7.value
        ).zfill(3)

        celula_m7.number_format = "@"

    # ======================================================
    # AUDITORIA
    # ======================================================

    ws_auditoria = wb.create_sheet(
        "AUDITORIA"
    )

    ws_auditoria.append([
        "Data",
        "Loteria",
        "Status",
        "Quantidade",
        "Novos",
        "URL",
        "Erro",
        "Processado em",
    ])

    for registro in auditoria:

        ws_auditoria.append([
            registro.get("data", ""),
            registro.get("loteria", ""),
            registro.get("status", ""),
            registro.get("quantidade", 0),
            registro.get("novos", 0),
            registro.get("url", ""),
            registro.get("erro", ""),
            registro.get("processado_em", ""),
        ])

    # ======================================================
    # AJUSTES VISUAIS
    # ======================================================

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:J{ws.max_row}"
    )

    larguras_resultados = {
        "A": 13,
        "B": 13,
        "C": 10,
        "D": 9,
        "E": 9,
        "F": 9,
        "G": 9,
        "H": 9,
        "I": 9,
        "J": 9,
    }

    for coluna, largura in (
        larguras_resultados.items()
    ):
        ws.column_dimensions[
            coluna
        ].width = largura

    ws_auditoria.freeze_panes = "A2"

    ws_auditoria.auto_filter.ref = (
        f"A1:H{ws_auditoria.max_row}"
    )

    larguras_auditoria = {
        "A": 13,
        "B": 13,
        "C": 24,
        "D": 13,
        "E": 10,
        "F": 75,
        "G": 40,
        "H": 22,
    }

    for coluna, largura in (
        larguras_auditoria.items()
    ):
        ws_auditoria.column_dimensions[
            coluna
        ].width = largura

    # ======================================================
    # SALVA NO PERSISTENT DISK
    # ======================================================

    wb.save(
        ARQUIVO_EXCEL
    )

    return {
        "resultados": len(
            resultados
        ),
        "auditoria": len(
            auditoria
        ),
        "arquivo": str(
            ARQUIVO_EXCEL
        ),
    }


# ==========================================================
# NOVA COLETA HISTÓRICA
# AVAL-PE + CAMINHO-DA-SORTE + MINAS-MG
# ==========================================================

def novas_estado_inicial():
    return {
        "status": "nao_iniciada",
        "data_inicial": NOVAS_DATA_INICIAL.strftime("%d/%m/%Y"),
        "data_final": NOVAS_DATA_FINAL.strftime("%d/%m/%Y"),
        "data_atual": None,
        "loteria_atual": None,
        "paginas_processadas": 0,
        "resultados_coletados": 0,
        "falhas": 0,
        "progresso": 0.0,
        "mensagem": "A coleta das novas loterias ainda não foi iniciada.",
    }


def salvar_estado_novas(estado):
    temporario = ARQUIVO_NOVAS_ESTADO.with_suffix(".tmp")
    with open(temporario, "w", encoding="utf-8") as arquivo:
        json.dump(estado, arquivo, ensure_ascii=False, indent=2)
    os.replace(temporario, ARQUIVO_NOVAS_ESTADO)


def carregar_estado_novas():
    if not ARQUIVO_NOVAS_ESTADO.exists():
        estado = novas_estado_inicial()
        salvar_estado_novas(estado)
        return estado

    try:
        with open(ARQUIVO_NOVAS_ESTADO, "r", encoding="utf-8") as arquivo:
            return json.load(arquivo)
    except Exception:
        logging.exception("Erro ao carregar estado das novas loterias.")
        return novas_estado_inicial()


def carregar_chaves_novas():
    chaves = set()

    if not ARQUIVO_NOVAS_RESULTADOS.exists():
        return chaves

    with open(ARQUIVO_NOVAS_RESULTADOS, "r", encoding="utf-8") as arquivo:
        for linha in arquivo:
            linha = linha.strip()
            if not linha:
                continue
            try:
                registro = json.loads(linha)
            except Exception:
                continue

            chaves.add(
                f"{registro.get('data', '')}|"
                f"{registro.get('loteria', '')}|"
                f"{registro.get('horario', '')}"
            )

    return chaves


def carregar_paginas_novas_concluidas():
    concluidas = set()

    if not ARQUIVO_NOVAS_AUDITORIA.exists():
        return concluidas

    with open(ARQUIVO_NOVAS_AUDITORIA, "r", encoding="utf-8") as arquivo:
        for linha in arquivo:
            linha = linha.strip()
            if not linha:
                continue
            try:
                registro = json.loads(linha)
            except Exception:
                continue

            if registro.get("status") not in {
                "ok",
                "nao_encontrado",
                "sem_resultados_validos",
            }:
                continue

            concluidas.add(
                f"{registro.get('data', '')}|"
                f"{registro.get('loteria', '')}"
            )

    return concluidas


def montar_url_nova_loteria(loteria, data_obj):
    cfg = NOVAS_LOTERIAS[loteria]
    data_iso = data_obj.strftime("%Y-%m-%d")
    return cfg["url_base"] + cfg["slug"] + data_iso


def pertence_nova_loteria(loteria, nome_variavel):
    """
    As páginas consultadas já são específicas de cada conjunto
    de resultados no Resultado Fácil.

    Por isso, para estas três novas loterias, a validação principal é:
    1) a página/URL correta;
    2) o horário estar na lista permitida;
    3) o bloco NÃO ser Federal.

    Isso é importante porque o Resultado Fácil usa nomes diferentes
    dentro da mesma página. Exemplo MINAS-MG:
      12h -> ALVORADA - MG
      15h -> MINAS DIA - MG
      19h -> MINAS NOITE - MG
      21h -> PREFERIDA - MG

    AVAL-PE também pode usar denominações diferentes em alguns horários.
    """
    texto = normalizar_texto(nome_variavel).upper()

    if "FEDERAL" in texto:
        return False

    return loteria in NOVAS_LOTERIAS


def extrair_resultados_nova_loteria(html, loteria, data_obj, url):
    estrutura = extrair_dataset_resultadofacil(html)
    variaveis = estrutura.get("variaveis", [])

    cfg = NOVAS_LOTERIAS[loteria]
    horarios_origem = cfg["horarios_origem"]
    normalizacao = cfg["normalizacao_horarios"]

    sorteios = {}

    for item in variaveis:
        if not isinstance(item, dict):
            continue

        nome = obter_campo_item(item, ["name", "nome"])
        valor = obter_campo_item(item, ["value", "valor"])

        if not nome or not valor:
            continue

        # A Federal pode aparecer misturada nessas páginas.
        # Nunca deve ser incorporada às novas loterias.
        if parece_federal(nome):
            continue

        if not pertence_nova_loteria(loteria, nome):
            continue

        horario_origem = extrair_horario_variavel(nome)

        if horario_origem not in horarios_origem:
            continue

        horario_destino = normalizacao.get(
            horario_origem,
            horario_origem
        )

        posicao = extrair_posicao_premio(nome)
        if posicao is None:
            continue

        milhar = extrair_milhar_valor(valor)
        if not milhar:
            continue

        sorteios.setdefault(horario_destino, {})
        sorteios[horario_destino][posicao] = milhar

    resultados = []

    for horario in sorted(sorteios.keys()):
        premios_dict = sorteios[horario]

        if not all(p in premios_dict for p in range(1, 6)):
            continue

        premios = [
            premios_dict[1],
            premios_dict[2],
            premios_dict[3],
            premios_dict[4],
            premios_dict[5],
        ]

        m6, m7 = calcular_premios_6_7(premios)

        resultados.append({
            "data": data_obj.strftime("%d/%m/%Y"),
            "loteria": loteria,
            "horario": horario,
            "m1": premios[0],
            "m2": premios[1],
            "m3": premios[2],
            "m4": premios[3],
            "m5": premios[4],
            "m6": m6,
            "m7": m7,
            "url": url,
            "origem": "RESULTADO_FACIL",
        })

    return resultados


def buscar_nova_loteria_data(loteria, data_obj):
    url = montar_url_nova_loteria(loteria, data_obj)

    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
    except Exception as e:
        return {
            "ok": False,
            "status": "erro_rede",
            "erro": str(e),
            "url": url,
            "resultados": [],
        }

    if resp.status_code == 404:
        return {
            "ok": True,
            "status": "nao_encontrado",
            "url": url,
            "resultados": [],
        }

    if resp.status_code == 403:
        return {
            "ok": False,
            "status": "bloqueado_403",
            "url": url,
            "resultados": [],
        }

    if resp.status_code != 200:
        return {
            "ok": False,
            "status": f"http_{resp.status_code}",
            "url": url,
            "resultados": [],
        }

    resultados = extrair_resultados_nova_loteria(
        resp.text,
        loteria,
        data_obj,
        url,
    )

    if not resultados:
        return {
            "ok": True,
            "status": "sem_resultados_validos",
            "url": url,
            "resultados": [],
        }

    return {
        "ok": True,
        "status": "ok",
        "url": url,
        "resultados": resultados,
    }


def executar_coleta_novas_loterias():
    global NOVAS_THREAD

    chaves_resultados = carregar_chaves_novas()
    paginas_concluidas = carregar_paginas_novas_concluidas()
    estado = carregar_estado_novas()

    estado.update({
        "status": "executando",
        "data_atual": None,
        "loteria_atual": None,
        "paginas_processadas": len(paginas_concluidas),
        "resultados_coletados": len(chaves_resultados),
        "falhas": 0,
        "progresso": round(
            (len(paginas_concluidas) / NOVAS_TOTAL_PAGINAS) * 100,
            2
        ),
        "mensagem": "Coleta histórica das novas loterias em andamento.",
    })

    salvar_estado_novas(estado)

    data_atual = NOVAS_DATA_INICIAL

    try:
        while data_atual <= NOVAS_DATA_FINAL:
            data_br = data_atual.strftime("%d/%m/%Y")

            for loteria in NOVAS_LOTERIAS:
                chave_pagina = f"{data_br}|{loteria}"

                if chave_pagina in paginas_concluidas:
                    continue

                estado["data_atual"] = data_br
                estado["loteria_atual"] = loteria
                estado["mensagem"] = f"Consultando {loteria} em {data_br}."
                salvar_estado_novas(estado)

                retorno = buscar_nova_loteria_data(
                    loteria,
                    data_atual
                )

                status = retorno.get("status", "desconhecido")
                resultados = retorno.get("resultados", [])
                novos = 0

                if status == "ok":
                    for resultado in resultados:
                        chave = (
                            f"{resultado['data']}|"
                            f"{resultado['loteria']}|"
                            f"{resultado['horario']}"
                        )

                        if chave in chaves_resultados:
                            continue

                        adicionar_jsonl(
                            ARQUIVO_NOVAS_RESULTADOS,
                            resultado
                        )
                        chaves_resultados.add(chave)
                        novos += 1

                adicionar_jsonl(
                    ARQUIVO_NOVAS_AUDITORIA,
                    {
                        "data": data_br,
                        "loteria": loteria,
                        "status": status,
                        "quantidade": len(resultados),
                        "novos": novos,
                        "url": retorno.get("url", ""),
                        "erro": retorno.get("erro", ""),
                        "processado_em": datetime.now().strftime(
                            "%d/%m/%Y %H:%M:%S"
                        ),
                    }
                )

                if status in {
                    "ok",
                    "nao_encontrado",
                    "sem_resultados_validos",
                }:
                    paginas_concluidas.add(chave_pagina)
                else:
                    estado["falhas"] += 1

                estado["paginas_processadas"] = len(paginas_concluidas)
                estado["resultados_coletados"] = len(chaves_resultados)
                estado["progresso"] = round(
                    (len(paginas_concluidas) / NOVAS_TOTAL_PAGINAS) * 100,
                    2
                )
                salvar_estado_novas(estado)

                time.sleep(0.8)

            data_atual += timedelta(days=1)

        estado.update({
            "status": "concluida",
            "data_atual": NOVAS_DATA_FINAL.strftime("%d/%m/%Y"),
            "loteria_atual": None,
            "paginas_processadas": len(paginas_concluidas),
            "resultados_coletados": len(chaves_resultados),
            "progresso": round(
                (len(paginas_concluidas) / NOVAS_TOTAL_PAGINAS) * 100,
                2
            ),
            "mensagem": "Coleta histórica das novas loterias concluída.",
        })
        salvar_estado_novas(estado)

    except Exception as e:
        logging.exception(
            "Falha geral na coleta das novas loterias."
        )
        estado.update({
            "status": "erro",
            "mensagem": str(e),
        })
        salvar_estado_novas(estado)

    finally:
        with NOVAS_LOCK:
            NOVAS_THREAD = None


def gerar_excel_novas_loterias():
    resultados = carregar_jsonl(
        ARQUIVO_NOVAS_RESULTADOS
    )
    auditoria = carregar_jsonl(
        ARQUIVO_NOVAS_AUDITORIA
    )

    if not resultados:
        raise ValueError(
            "Nenhum resultado das novas loterias foi coletado."
        )

    ordem = {
        "AVAL-PE": 1,
        "CAMINHO-DA-SORTE": 2,
        "MINAS-MG": 3,
    }

    resultados.sort(
        key=lambda item: (
            datetime.strptime(
                item.get("data", "01/01/1900"),
                "%d/%m/%Y"
            ),
            ordem.get(
                item.get("loteria", ""),
                999
            ),
            str(item.get("horario", "")).zfill(2),
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "RESULTADOS"

    ws.append([
        "Data",
        "Loteria",
        "Horário",
        "M1",
        "M2",
        "M3",
        "M4",
        "M5",
        "M6",
        "M7",
    ])

    for resultado in resultados:
        ws.append([
            resultado.get("data", ""),
            resultado.get("loteria", ""),
            resultado.get("horario", ""),
            resultado.get("m1", ""),
            resultado.get("m2", ""),
            resultado.get("m3", ""),
            resultado.get("m4", ""),
            resultado.get("m5", ""),
            resultado.get("m6", ""),
            resultado.get("m7", ""),
        ])

    for linha in range(2, ws.max_row + 1):
        ws.cell(linha, 3).number_format = "@"

        for coluna in range(4, 10):
            celula = ws.cell(linha, coluna)
            celula.value = str(celula.value).zfill(4)
            celula.number_format = "@"

        celula_m7 = ws.cell(linha, 10)
        celula_m7.value = str(celula_m7.value).zfill(3)
        celula_m7.number_format = "@"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    for coluna, largura in {
        "A": 13,
        "B": 24,
        "C": 10,
        "D": 9,
        "E": 9,
        "F": 9,
        "G": 9,
        "H": 9,
        "I": 9,
        "J": 9,
    }.items():
        ws.column_dimensions[coluna].width = largura

    ws_auditoria = wb.create_sheet("AUDITORIA")
    ws_auditoria.append([
        "Data",
        "Loteria",
        "Status",
        "Quantidade",
        "Novos",
        "URL",
        "Erro",
        "Processado em",
    ])

    for registro in auditoria:
        ws_auditoria.append([
            registro.get("data", ""),
            registro.get("loteria", ""),
            registro.get("status", ""),
            registro.get("quantidade", 0),
            registro.get("novos", 0),
            registro.get("url", ""),
            registro.get("erro", ""),
            registro.get("processado_em", ""),
        ])

    wb.save(ARQUIVO_NOVAS_EXCEL)

    return {
        "resultados": len(resultados),
        "auditoria": len(auditoria),
        "arquivo": str(ARQUIVO_NOVAS_EXCEL),
    }


# ==========================================================
# ROTAS DAS NOVAS LOTERIAS
# ==========================================================

@app.route("/novas-loterias/teste/<loteria>/<data_teste>")
def teste_nova_loteria(loteria, data_teste):
    loteria = loteria.upper()

    if loteria not in NOVAS_LOTERIAS:
        return jsonify({
            "ok": False,
            "erro": "Loteria inválida.",
            "loterias": list(NOVAS_LOTERIAS.keys()),
        }), 400

    try:
        data_obj = datetime.strptime(
            data_teste,
            "%Y-%m-%d"
        )
    except ValueError:
        return jsonify({
            "ok": False,
            "erro": "Data inválida. Use YYYY-MM-DD.",
        }), 400

    retorno = buscar_nova_loteria_data(
        loteria,
        data_obj
    )

    return jsonify({
        "ok": True,
        "data_consulta": data_teste,
        "loteria": loteria,
        **retorno,
    })


@app.route("/novas-loterias/debug/<loteria>/<data_teste>")
def debug_nova_loteria(loteria, data_teste):
    loteria = loteria.upper()

    if loteria not in NOVAS_LOTERIAS:
        return jsonify({
            "ok": False,
            "erro": "Loteria inválida.",
        }), 400

    try:
        data_obj = datetime.strptime(
            data_teste,
            "%Y-%m-%d"
        )
    except ValueError:
        return jsonify({
            "ok": False,
            "erro": "Data inválida. Use YYYY-MM-DD.",
        }), 400

    url = montar_url_nova_loteria(
        loteria,
        data_obj
    )

    resp = requests.get(
        url,
        headers=HEADERS,
        timeout=30,
    )

    estrutura = extrair_dataset_resultadofacil(
        resp.text
    )

    candidatos = []

    for item in estrutura.get("variaveis", []):
        if not isinstance(item, dict):
            continue

        nome = obter_campo_item(
            item,
            ["name", "nome"]
        )
        valor = obter_campo_item(
            item,
            ["value", "valor"]
        )
        horario = extrair_horario_variavel(
            nome
        )

        candidatos.append({
            "nome": nome,
            "valor": valor,
            "horario_origem": horario,
            "horario_destino": (
                NOVAS_LOTERIAS[
                    loteria
                ]["normalizacao_horarios"].get(
                    horario,
                    horario
                )
            ),
            "posicao": extrair_posicao_premio(
                nome
            ),
            "federal": parece_federal(
                nome
            ),
            "aceito_banca": pertence_nova_loteria(
                loteria,
                nome
            ),
            "horario_valido": (
                horario
                in NOVAS_LOTERIAS[
                    loteria
                ]["horarios_origem"]
            ),
        })

    return jsonify({
        "ok": True,
        "status_http": resp.status_code,
        "url": url,
        "total_variaveis_dataset": len(
            estrutura.get("variaveis", [])
        ),
        "candidatos": candidatos,
    })


@app.route("/novas-loterias/status")
def status_novas_loterias():
    estado = carregar_estado_novas()

    return jsonify({
        "ok": True,
        "disco": str(DIRETORIO_DADOS),
        "total_dias_previstos": NOVAS_TOTAL_DIAS,
        "total_paginas_previstas": NOVAS_TOTAL_PAGINAS,
        "estado": estado,
        "arquivos": {
            "estado": ARQUIVO_NOVAS_ESTADO.exists(),
            "resultados": ARQUIVO_NOVAS_RESULTADOS.exists(),
            "auditoria": ARQUIVO_NOVAS_AUDITORIA.exists(),
            "excel": ARQUIVO_NOVAS_EXCEL.exists(),
        },
    })


@app.route("/novas-loterias/iniciar")
def iniciar_novas_loterias():
    global NOVAS_THREAD

    with NOVAS_LOCK:
        if (
            NOVAS_THREAD is not None
            and NOVAS_THREAD.is_alive()
        ):
            return jsonify({
                "ok": False,
                "mensagem": (
                    "A coleta das novas loterias já está em execução."
                ),
                "estado": carregar_estado_novas(),
            }), 409

        NOVAS_THREAD = threading.Thread(
            target=executar_coleta_novas_loterias,
            name="coleta_novas_loterias",
            daemon=True,
        )
        NOVAS_THREAD.start()

    return jsonify({
        "ok": True,
        "mensagem": (
            "Coleta histórica das novas loterias iniciada."
        ),
        "periodo": {
            "inicio": NOVAS_DATA_INICIAL.strftime("%d/%m/%Y"),
            "fim": NOVAS_DATA_FINAL.strftime("%d/%m/%Y"),
        },
        "loterias": list(NOVAS_LOTERIAS.keys()),
        "total_paginas": NOVAS_TOTAL_PAGINAS,
        "status_url": "/novas-loterias/status",
    })


@app.route("/novas-loterias/gerar-excel")
def rota_gerar_excel_novas_loterias():
    estado = carregar_estado_novas()

    if estado.get("status") != "concluida":
        return jsonify({
            "ok": False,
            "mensagem": (
                "A coleta das novas loterias ainda não foi concluída."
            ),
            "estado": estado,
        }), 409

    try:
        resumo = gerar_excel_novas_loterias()

        return jsonify({
            "ok": True,
            "mensagem": (
                "Planilha das novas loterias gerada com sucesso."
            ),
            **resumo,
            "download": "/novas-loterias/baixar",
        })
    except Exception as e:
        logging.exception(
            "Erro ao gerar Excel das novas loterias."
        )
        return jsonify({
            "ok": False,
            "erro": str(e),
        }), 500


@app.route("/novas-loterias/baixar")
def baixar_excel_novas_loterias():
    if not ARQUIVO_NOVAS_EXCEL.exists():
        return jsonify({
            "ok": False,
            "mensagem": (
                "A planilha das novas loterias ainda não foi gerada. "
                "Acesse /novas-loterias/gerar-excel primeiro."
            ),
        }), 404

    return send_file(
        ARQUIVO_NOVAS_EXCEL,
        as_attachment=True,
        download_name=(
            "aval_pe_caminho_da_sorte_minas_mg_"
            "01-01-2025_a_14-08-2026.xlsx"
        ),
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


# ==========================================================
# DIAGNÓSTICO HISTÓRICO JB CERTO - LOTEP 20H
# ==========================================================

def diagnosticar_historico_lotep_20_jbcerto():

    resp = requests.get(
        URL_JBCERTO_LOTEP,
        headers=HEADERS,
        timeout=30,
    )

    resp.raise_for_status()

    soup = BeautifulSoup(
        resp.text,
        "html.parser"
    )

    resultados = []

    for tabela in soup.find_all("table"):

        # ==================================================
        # PROCURA O TÍTULO DO SORTEIO ANTES DA TABELA
        # ==================================================

        titulo = ""

        elemento = tabela.find_previous()

        limite = 0

        while elemento is not None and limite < 40:
            limite += 1

            try:
                texto = normalizar_texto(
                    elemento.get_text(
                        " ",
                        strip=True
                    )
                )
            except Exception:
                elemento = elemento.find_previous()
                continue

            texto_lower = texto.lower()

            if (
                "lotep" in texto_lower
                or "paratodos" in texto_lower
            ):
                if re.search(
                    r"\b20\s*(?:h|horas?)\b",
                    texto_lower
                ):
                    titulo = texto
                    break

            elemento = elemento.find_previous()

        if not titulo:
            continue

        # ==================================================
        # PROCURA A DATA ANTERIOR À TABELA
        # ==================================================

        data_br = ""

        elemento = tabela.find_previous()

        limite = 0

        while elemento is not None and limite < 30:
            limite += 1

            try:
                texto = normalizar_texto(
                    elemento.get_text(
                        " ",
                        strip=True
                    )
                )
            except Exception:
                elemento = elemento.find_previous()
                continue

            match_data = re.search(
                r"\b(\d{2}/\d{2}/\d{4})\b",
                texto
            )

            if match_data:
                data_br = match_data.group(1)
                break

            elemento = elemento.find_previous()

        if not data_br:
            continue

        # ==================================================
        # EXTRAI M1 ATÉ M5
        # ==================================================

        premios = {}

        for linha in tabela.find_all("tr"):

            colunas = linha.find_all(
                ["td", "th"]
            )

            if len(colunas) < 2:
                continue

            premio_txt = normalizar_texto(
                colunas[0].get_text(
                    " ",
                    strip=True
                )
            )

            milhar_txt = normalizar_texto(
                colunas[1].get_text(
                    " ",
                    strip=True
                )
            )

            match_premio = re.search(
                r"([1-5])",
                premio_txt
            )

            if not match_premio:
                continue

            posicao = int(
                match_premio.group(1)
            )

            milhar = formatar_milhar(
                milhar_txt
            )

            if not milhar:
                continue

            premios[
                posicao
            ] = milhar

        if not all(
            p in premios
            for p in range(1, 6)
        ):
            continue

        lista_premios = [
            premios[1],
            premios[2],
            premios[3],
            premios[4],
            premios[5],
        ]

        m6, m7 = calcular_premios_6_7(
            lista_premios
        )

        resultados.append({
            "data": data_br,
            "loteria": "LOTEP",
            "horario": "20",
            "m1": premios[1],
            "m2": premios[2],
            "m3": premios[3],
            "m4": premios[4],
            "m5": premios[5],
            "m6": m6,
            "m7": m7,
            "titulo": titulo,
        })

    # ======================================================
    # REMOVE DUPLICIDADES
    # ======================================================

    unicos = {}

    for resultado in resultados:

        chave = (
            resultado["data"],
            resultado["horario"],
        )

        unicos[chave] = resultado

    resultados = list(
        unicos.values()
    )

    resultados.sort(
        key=lambda item: datetime.strptime(
            item["data"],
            "%d/%m/%Y"
        )
    )

    return {
        "url": URL_JBCERTO_LOTEP,
        "total": len(resultados),
        "primeira_data": (
            resultados[0]["data"]
            if resultados
            else None
        ),
        "ultima_data": (
            resultados[-1]["data"]
            if resultados
            else None
        ),
        "resultados": resultados,
    }


# ==========================================================
# COLETA HISTÓRICA COMPLEMENTAR - LOTEP 20H
# ==========================================================

def lotep20_estado_inicial():
    return {
        "status": "nao_iniciada",
        "data_inicial": LOTEP20_DATA_INICIAL.strftime("%d/%m/%Y"),
        "data_final": LOTEP20_DATA_FINAL.strftime("%d/%m/%Y"),
        "data_atual": None,
        "dias_processados": 0,
        "resultados_coletados": 0,
        "falhas": 0,
        "progresso": 0.0,
        "mensagem": "A coleta LOTEP 20h ainda não foi iniciada.",
    }


def salvar_estado_lotep20(estado):
    temporario = ARQUIVO_LOTEP20_ESTADO.with_suffix(".tmp")

    with open(temporario, "w", encoding="utf-8") as arquivo:
        json.dump(
            estado,
            arquivo,
            ensure_ascii=False,
            indent=2,
        )

    os.replace(temporario, ARQUIVO_LOTEP20_ESTADO)


def carregar_estado_lotep20():
    if not ARQUIVO_LOTEP20_ESTADO.exists():
        estado = lotep20_estado_inicial()
        salvar_estado_lotep20(estado)
        return estado

    try:
        with open(
            ARQUIVO_LOTEP20_ESTADO,
            "r",
            encoding="utf-8"
        ) as arquivo:
            return json.load(arquivo)

    except Exception:
        logging.exception("Erro ao carregar estado LOTEP 20h.")
        return lotep20_estado_inicial()


def carregar_datas_lotep20_concluidas():
    concluidas = set()

    if not ARQUIVO_LOTEP20_AUDITORIA.exists():
        return concluidas

    try:
        with open(
            ARQUIVO_LOTEP20_AUDITORIA,
            "r",
            encoding="utf-8"
        ) as arquivo:
            for linha in arquivo:
                linha = linha.strip()

                if not linha:
                    continue

                try:
                    registro = json.loads(linha)
                except Exception:
                    continue

                if registro.get("status") not in {
                    "ok",
                    "nao_encontrado",
                    "sem_lotep20",
                }:
                    continue

                data = registro.get("data", "")

                if data:
                    concluidas.add(data)

    except Exception:
        logging.exception("Erro ao carregar auditoria LOTEP 20h.")

    return concluidas


def carregar_chaves_lotep20():
    chaves = set()

    if not ARQUIVO_LOTEP20_RESULTADOS.exists():
        return chaves

    try:
        with open(
            ARQUIVO_LOTEP20_RESULTADOS,
            "r",
            encoding="utf-8"
        ) as arquivo:
            for linha in arquivo:
                linha = linha.strip()

                if not linha:
                    continue

                try:
                    registro = json.loads(linha)
                except Exception:
                    continue

                chave = (
                    f"{registro.get('data', '')}|"
                    f"{registro.get('loteria', '')}|"
                    f"{registro.get('horario', '')}"
                )

                chaves.add(chave)

    except Exception:
        logging.exception("Erro ao carregar resultados LOTEP 20h.")

    return chaves


def nome_dia_semana_lotep20(data_obj):
    nomes = {
        0: "segunda-feira",
        1: "terca-feira",
        2: "quarta-feira",
        3: "quinta-feira",
        4: "sexta-feira",
        5: "sabado",
        6: "domingo",
    }

    return nomes[data_obj.weekday()]


def montar_url_lotep20(data_obj):
    """
    Página histórica da PARATODOS PB.

    Exemplo:
    12/08/2026 ->
    https://www.resultadofacil.com.br/resultados-paratodos-pb-do-dia-2026-08-12
    """
    data_iso = data_obj.strftime("%Y-%m-%d")

    return (
        f"{LOTEP20_URL_BASE}"
        f"resultados-paratodos-pb-do-dia-"
        f"{data_iso}"
    )


def _normalizar_sem_acentos_lotep20(texto):
    texto = normalizar_texto(texto).lower()

    trocas = {
        "á": "a",
        "à": "a",
        "ã": "a",
        "â": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ç": "c",
    }

    for origem, destino in trocas.items():
        texto = texto.replace(origem, destino)

    return texto


def _texto_tem_lotep20(texto):
    texto = _normalizar_sem_acentos_lotep20(texto)

    return (
        "paratodos" in texto
        and re.search(
            r"\b20\s*(?:h|horas?)\b",
            texto
        ) is not None
    )


def _extrair_premios_de_tabela_lotep20(tabela):
    premios = {}

    for linha in tabela.find_all("tr"):
        colunas = linha.find_all(["td", "th"])

        if len(colunas) < 2:
            continue

        texto_primeira = normalizar_texto(
            colunas[0].get_text(" ", strip=True)
        )

        match_premio = re.search(
            r"\b([1-5])\s*[ºoªa]?\b",
            texto_primeira,
            flags=re.IGNORECASE,
        )

        if not match_premio:
            continue

        posicao = int(match_premio.group(1))
        milhar = ""

        for coluna in colunas[1:]:
            texto_coluna = normalizar_texto(
                coluna.get_text(" ", strip=True)
            )

            match_milhar = re.search(
                r"(?<!\d)(\d{1,4})(?!\d)",
                texto_coluna
            )

            if match_milhar:
                milhar = match_milhar.group(1).zfill(4)
                break

        if milhar:
            premios[posicao] = milhar

    if not all(p in premios for p in range(1, 6)):
        return []

    return [
        premios[1],
        premios[2],
        premios[3],
        premios[4],
        premios[5],
    ]


def nome_variavel_eh_paratodos_pb(nome):
    texto = _normalizar_sem_acentos_lotep20(nome).upper()

    if "FEDERAL" in texto:
        return False

    return (
        "PARATODOS" in texto
        or "PARA TODOS" in texto
        or " PB " in f" {texto} "
        or "PARAIBA" in texto
    )


def extrair_paratodos_pb_resultadofacil(html, data_obj, url):
    """
    Extrai os sorteios 09h/09:45 e 20h da PARATODOS PB
    e normaliza ambos para a loteria LOTEP do LoteriasDB.
    """
    estrutura = extrair_dataset_resultadofacil(html)
    variaveis = estrutura.get("variaveis", [])

    sorteios = {
        "09": {},
        "20": {},
    }

    nomes_origem = {
        "09": [],
        "20": [],
    }

    for item in variaveis:
        if not isinstance(item, dict):
            continue

        nome = obter_campo_item(item, ["name", "nome"])
        valor = obter_campo_item(item, ["value", "valor"])

        if not nome or not valor:
            continue

        if not nome_variavel_eh_paratodos_pb(nome):
            continue

        horario = extrair_horario_variavel(nome)

        if horario not in {"09", "20"}:
            continue

        posicao = extrair_posicao_premio(nome)

        if posicao is None:
            continue

        milhar = extrair_milhar_valor(valor)

        if not milhar:
            continue

        sorteios[horario][posicao] = milhar
        nomes_origem[horario].append(nome)

    resultados = []

    for horario in ["09", "20"]:
        premios = sorteios[horario]

        if not all(p in premios for p in range(1, 6)):
            continue

        lista = [
            premios[1],
            premios[2],
            premios[3],
            premios[4],
            premios[5],
        ]

        m6, m7 = calcular_premios_6_7(lista)

        resultados.append({
            "data": data_obj.strftime("%d/%m/%Y"),
            "loteria": "LOTEP",
            "horario": horario,
            "m1": premios[1],
            "m2": premios[2],
            "m3": premios[3],
            "m4": premios[4],
            "m5": premios[5],
            "m6": m6,
            "m7": m7,
            "url": url,
            "origem": "RESULTADO_FACIL_PARATODOS_PB",
            "variaveis_origem": nomes_origem[horario],
        })

    return resultados


def buscar_lotep20_data(data_obj):
    url = montar_url_lotep20(data_obj)

    try:
        resp = requests.get(
            url,
            headers=HEADERS,
            timeout=30,
        )
    except Exception as e:
        return {
            "status": "erro_rede",
            "resultados": [],
            "url": url,
            "erro": str(e),
        }

    if resp.status_code == 404:
        return {
            "status": "nao_encontrado",
            "resultados": [],
            "url": url,
            "erro": "",
        }

    if resp.status_code == 403:
        return {
            "status": "bloqueado_403",
            "resultados": [],
            "url": url,
            "erro": "",
        }

    if resp.status_code != 200:
        return {
            "status": f"http_{resp.status_code}",
            "resultados": [],
            "url": url,
            "erro": "",
        }

    resultados = extrair_paratodos_pb_resultadofacil(
        resp.text,
        data_obj,
        url,
    )

    if not resultados:
        return {
            "status": "sem_paratodos_pb",
            "resultados": [],
            "url": url,
            "erro": "",
        }

    return {
        "status": "ok",
        "resultados": resultados,
        "url": url,
        "erro": "",
    }


def executar_coleta_lotep20():
    global LOTEP20_THREAD

    logging.info(
        "Iniciando coleta histórica PARATODOS PB 09h/20h -> LOTEP."
    )

    datas_concluidas = carregar_datas_lotep20_concluidas()
    chaves_resultados = carregar_chaves_lotep20()
    estado = carregar_estado_lotep20()

    estado.update({
        "status": "executando",
        "data_atual": None,
        "dias_processados": len(datas_concluidas),
        "resultados_coletados": len(chaves_resultados),
        "falhas": 0,
        "progresso": round(
            (len(datas_concluidas) / LOTEP20_TOTAL_DIAS) * 100,
            2
        ),
        "mensagem": "Coleta histórica PARATODOS PB 09h/20h em andamento.",
    })

    salvar_estado_lotep20(estado)

    data_atual = LOTEP20_DATA_INICIAL

    try:
        while data_atual <= LOTEP20_DATA_FINAL:
            data_br = data_atual.strftime("%d/%m/%Y")

            if data_br in datas_concluidas:
                data_atual += timedelta(days=1)
                continue

            estado["data_atual"] = data_br
            estado["mensagem"] = f"Consultando PARATODOS PB em {data_br}."
            salvar_estado_lotep20(estado)

            retorno = buscar_lotep20_data(data_atual)
            status = retorno.get("status", "desconhecido")
            resultados = retorno.get("resultados", [])
            novos = 0

            if status == "ok":
                for resultado in resultados:
                    chave = (
                        f"{resultado['data']}|"
                        f"{resultado['loteria']}|"
                        f"{resultado['horario']}"
                    )

                    if chave in chaves_resultados:
                        continue

                    adicionar_jsonl(
                        ARQUIVO_LOTEP20_RESULTADOS,
                        resultado
                    )

                    chaves_resultados.add(chave)
                    novos += 1

            adicionar_jsonl(
                ARQUIVO_LOTEP20_AUDITORIA,
                {
                    "data": data_br,
                    "loteria": "LOTEP",
                    "horarios": [r.get("horario") for r in resultados],
                    "status": status,
                    "quantidade": len(resultados),
                    "novos": novos,
                    "url": retorno.get("url", ""),
                    "erro": retorno.get("erro", ""),
                    "processado_em": datetime.now().strftime(
                        "%d/%m/%Y %H:%M:%S"
                    ),
                }
            )

            if status in {
                "ok",
                "nao_encontrado",
                "sem_paratodos_pb",
            }:
                datas_concluidas.add(data_br)
            else:
                estado["falhas"] += 1

            estado["dias_processados"] = len(datas_concluidas)
            estado["resultados_coletados"] = len(chaves_resultados)
            estado["progresso"] = round(
                (len(datas_concluidas) / LOTEP20_TOTAL_DIAS) * 100,
                2
            )

            salvar_estado_lotep20(estado)

            time.sleep(0.8)
            data_atual += timedelta(days=1)

        estado.update({
            "status": "concluida",
            "data_atual": LOTEP20_DATA_FINAL.strftime("%d/%m/%Y"),
            "dias_processados": len(datas_concluidas),
            "resultados_coletados": len(chaves_resultados),
            "progresso": round(
                (len(datas_concluidas) / LOTEP20_TOTAL_DIAS) * 100,
                2
            ),
            "mensagem": "Coleta histórica PARATODOS PB 09h/20h concluída.",
        })

        salvar_estado_lotep20(estado)

    except Exception as e:
        logging.exception(
            "Falha geral na coleta PARATODOS PB 09h/20h."
        )

        estado.update({
            "status": "erro",
            "mensagem": str(e),
        })

        salvar_estado_lotep20(estado)

    finally:
        with LOTEP20_LOCK:
            LOTEP20_THREAD = None


def gerar_excel_lotep20():
    resultados = carregar_jsonl(ARQUIVO_LOTEP20_RESULTADOS)
    auditoria = carregar_jsonl(ARQUIVO_LOTEP20_AUDITORIA)

    if not resultados:
        raise ValueError(
            "Nenhum resultado PARATODOS PB 09h/20h foi coletado."
        )

    resultados.sort(
        key=lambda item: datetime.strptime(
            item.get("data", "01/01/1900"),
            "%d/%m/%Y"
        )
    )

    wb = Workbook()

    ws = wb.active
    ws.title = "LOTEP 09H E 20H"

    ws.append([
        "Data",
        "Loteria",
        "Horário",
        "M1",
        "M2",
        "M3",
        "M4",
        "M5",
        "M6",
        "M7",
        "Origem",
        "URL",
    ])

    for resultado in resultados:
        ws.append([
            resultado.get("data", ""),
            resultado.get("loteria", "LOTEP"),
            resultado.get("horario", "20"),
            resultado.get("m1", ""),
            resultado.get("m2", ""),
            resultado.get("m3", ""),
            resultado.get("m4", ""),
            resultado.get("m5", ""),
            resultado.get("m6", ""),
            resultado.get("m7", ""),
            resultado.get("origem", ""),
            resultado.get("url", ""),
        ])

    for linha in range(2, ws.max_row + 1):
        ws.cell(linha, 3).number_format = "@"

        for coluna in range(4, 10):
            celula = ws.cell(linha, coluna)
            celula.value = str(celula.value).zfill(4)
            celula.number_format = "@"

        celula_m7 = ws.cell(linha, 10)
        celula_m7.value = str(celula_m7.value).zfill(3)
        celula_m7.number_format = "@"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"

    larguras = {
        "A": 13,
        "B": 12,
        "C": 10,
        "D": 9,
        "E": 9,
        "F": 9,
        "G": 9,
        "H": 9,
        "I": 9,
        "J": 9,
        "K": 22,
        "L": 85,
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    ws_auditoria = wb.create_sheet("AUDITORIA")

    ws_auditoria.append([
        "Data",
        "Loteria",
        "Horário",
        "Status",
        "Quantidade",
        "Novos",
        "URL",
        "Erro",
        "Processado em",
    ])

    for registro in auditoria:
        ws_auditoria.append([
            registro.get("data", ""),
            registro.get("loteria", "LOTEP"),
            registro.get("horario", "20"),
            registro.get("status", ""),
            registro.get("quantidade", 0),
            registro.get("novos", 0),
            registro.get("url", ""),
            registro.get("erro", ""),
            registro.get("processado_em", ""),
        ])

    ws_auditoria.freeze_panes = "A2"
    ws_auditoria.auto_filter.ref = (
        f"A1:I{ws_auditoria.max_row}"
    )

    ws_auditoria.column_dimensions["A"].width = 13
    ws_auditoria.column_dimensions["B"].width = 12
    ws_auditoria.column_dimensions["C"].width = 10
    ws_auditoria.column_dimensions["D"].width = 22
    ws_auditoria.column_dimensions["E"].width = 12
    ws_auditoria.column_dimensions["F"].width = 10
    ws_auditoria.column_dimensions["G"].width = 85
    ws_auditoria.column_dimensions["H"].width = 45
    ws_auditoria.column_dimensions["I"].width = 22

    wb.save(ARQUIVO_LOTEP20_EXCEL)

    return {
        "resultados": len(resultados),
        "auditoria": len(auditoria),
        "arquivo": str(ARQUIVO_LOTEP20_EXCEL),
    }



# ==========================================================
# COLETA HISTÓRICA COMPLEMENTAR - PT-SP 15H
# BANDEIRANTES 15H -> PT-SP | 15
# ==========================================================

def ptsp15_estado_inicial():
    return {
        "status": "nao_iniciada",
        "data_inicial": PTSP15_DATA_INICIAL.strftime("%d/%m/%Y"),
        "data_final": PTSP15_DATA_FINAL.strftime("%d/%m/%Y"),
        "data_atual": None,
        "dias_processados": 0,
        "resultados_coletados": 0,
        "falhas": 0,
        "progresso": 0.0,
        "mensagem": "A coleta PT-SP 15h ainda não foi iniciada.",
    }


def salvar_estado_ptsp15(estado):
    temporario = ARQUIVO_PTSP15_ESTADO.with_suffix(".tmp")
    with open(temporario, "w", encoding="utf-8") as arquivo:
        json.dump(estado, arquivo, ensure_ascii=False, indent=2)
    os.replace(temporario, ARQUIVO_PTSP15_ESTADO)


def carregar_estado_ptsp15():
    if not ARQUIVO_PTSP15_ESTADO.exists():
        estado = ptsp15_estado_inicial()
        salvar_estado_ptsp15(estado)
        return estado

    try:
        with open(ARQUIVO_PTSP15_ESTADO, "r", encoding="utf-8") as arquivo:
            return json.load(arquivo)
    except Exception:
        logging.exception("Erro ao carregar estado PT-SP 15h.")
        return ptsp15_estado_inicial()


def carregar_datas_ptsp15_concluidas():
    concluidas = set()

    if not ARQUIVO_PTSP15_AUDITORIA.exists():
        return concluidas

    with open(ARQUIVO_PTSP15_AUDITORIA, "r", encoding="utf-8") as arquivo:
        for linha in arquivo:
            linha = linha.strip()
            if not linha:
                continue

            try:
                registro = json.loads(linha)
            except Exception:
                continue

            if registro.get("status") not in {
                "ok",
                "nao_encontrado",
                "sem_ptsp15",
            }:
                continue

            data = registro.get("data", "")
            if data:
                concluidas.add(data)

    return concluidas


def carregar_chaves_ptsp15():
    chaves = set()

    if not ARQUIVO_PTSP15_RESULTADOS.exists():
        return chaves

    with open(ARQUIVO_PTSP15_RESULTADOS, "r", encoding="utf-8") as arquivo:
        for linha in arquivo:
            linha = linha.strip()
            if not linha:
                continue

            try:
                registro = json.loads(linha)
            except Exception:
                continue

            chaves.add(
                f"{registro.get('data', '')}|"
                f"{registro.get('loteria', '')}|"
                f"{registro.get('horario', '')}"
            )

    return chaves


def montar_url_ptsp15(data_obj):
    data_iso = data_obj.strftime("%Y-%m-%d")

    return (
        f"{PTSP15_URL_BASE}"
        f"resultados-bandeirantes-do-dia-"
        f"{data_iso}"
    )


def _normalizar_sem_acentos_ptsp15(texto):
    texto = normalizar_texto(texto).lower()

    trocas = {
        "á": "a", "à": "a", "ã": "a", "â": "a",
        "é": "e", "ê": "e", "í": "i",
        "ó": "o", "ô": "o", "õ": "o",
        "ú": "u", "ç": "c",
    }

    for origem, destino in trocas.items():
        texto = texto.replace(origem, destino)

    return texto


def nome_variavel_eh_bandeirantes_15(nome):
    texto = _normalizar_sem_acentos_ptsp15(nome).upper()

    if "FEDERAL" in texto:
        return False

    if extrair_horario_variavel(nome) != "15":
        return False

    return (
        "BANDEIRANTES" in texto
        or "BANDEIRANTE" in texto
        or "SP" in texto
    )


def extrair_ptsp15_resultadofacil(html, data_obj, url):
    estrutura = extrair_dataset_resultadofacil(html)
    variaveis = estrutura.get("variaveis", [])

    premios = {}
    nomes_origem = []

    for item in variaveis:
        if not isinstance(item, dict):
            continue

        nome = obter_campo_item(item, ["name", "nome"])
        valor = obter_campo_item(item, ["value", "valor"])

        if not nome or not valor:
            continue

        if not nome_variavel_eh_bandeirantes_15(nome):
            continue

        posicao = extrair_posicao_premio(nome)
        if posicao is None:
            continue

        milhar = extrair_milhar_valor(valor)
        if not milhar:
            continue

        premios[posicao] = milhar
        nomes_origem.append(nome)

    if not all(p in premios for p in range(1, 6)):
        return None

    lista = [
        premios[1],
        premios[2],
        premios[3],
        premios[4],
        premios[5],
    ]

    m6, m7 = calcular_premios_6_7(lista)

    return {
        "data": data_obj.strftime("%d/%m/%Y"),
        "loteria": "PT-SP",
        "horario": "15",
        "m1": premios[1],
        "m2": premios[2],
        "m3": premios[3],
        "m4": premios[4],
        "m5": premios[5],
        "m6": m6,
        "m7": m7,
        "url": url,
        "origem": "RESULTADO_FACIL_BANDEIRANTES",
        "variaveis_origem": nomes_origem,
    }


def buscar_ptsp15_data(data_obj):
    url = montar_url_ptsp15(data_obj)

    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
    except Exception as e:
        return {
            "status": "erro_rede",
            "resultado": None,
            "url": url,
            "erro": str(e),
        }

    if resp.status_code == 404:
        return {
            "status": "nao_encontrado",
            "resultado": None,
            "url": url,
            "erro": "",
        }

    if resp.status_code == 403:
        return {
            "status": "bloqueado_403",
            "resultado": None,
            "url": url,
            "erro": "",
        }

    if resp.status_code != 200:
        return {
            "status": f"http_{resp.status_code}",
            "resultado": None,
            "url": url,
            "erro": "",
        }

    resultado = extrair_ptsp15_resultadofacil(
        resp.text,
        data_obj,
        url,
    )

    if resultado is None:
        return {
            "status": "sem_ptsp15",
            "resultado": None,
            "url": url,
            "erro": "",
        }

    return {
        "status": "ok",
        "resultado": resultado,
        "url": url,
        "erro": "",
    }


def executar_coleta_ptsp15():
    global PTSP15_THREAD

    datas_concluidas = carregar_datas_ptsp15_concluidas()
    chaves_resultados = carregar_chaves_ptsp15()
    estado = carregar_estado_ptsp15()

    estado.update({
        "status": "executando",
        "data_atual": None,
        "dias_processados": len(datas_concluidas),
        "resultados_coletados": len(chaves_resultados),
        "falhas": 0,
        "progresso": round(
            (len(datas_concluidas) / PTSP15_TOTAL_DIAS) * 100,
            2
        ),
        "mensagem": "Coleta histórica PT-SP 15h em andamento.",
    })

    salvar_estado_ptsp15(estado)

    data_atual = PTSP15_DATA_INICIAL

    try:
        while data_atual <= PTSP15_DATA_FINAL:
            data_br = data_atual.strftime("%d/%m/%Y")

            if data_br in datas_concluidas:
                data_atual += timedelta(days=1)
                continue

            estado["data_atual"] = data_br
            estado["mensagem"] = (
                f"Consultando Bandeirantes 15h em {data_br}."
            )
            salvar_estado_ptsp15(estado)

            retorno = buscar_ptsp15_data(data_atual)
            status = retorno.get("status", "desconhecido")
            resultado = retorno.get("resultado")
            novo = 0

            if status == "ok" and resultado:
                chave = (
                    f"{resultado['data']}|"
                    f"{resultado['loteria']}|"
                    f"{resultado['horario']}"
                )

                if chave not in chaves_resultados:
                    adicionar_jsonl(
                        ARQUIVO_PTSP15_RESULTADOS,
                        resultado
                    )
                    chaves_resultados.add(chave)
                    novo = 1

            adicionar_jsonl(
                ARQUIVO_PTSP15_AUDITORIA,
                {
                    "data": data_br,
                    "loteria": "PT-SP",
                    "horario": "15",
                    "status": status,
                    "quantidade": 1 if resultado else 0,
                    "novos": novo,
                    "url": retorno.get("url", ""),
                    "erro": retorno.get("erro", ""),
                    "processado_em": datetime.now().strftime(
                        "%d/%m/%Y %H:%M:%S"
                    ),
                }
            )

            if status in {
                "ok",
                "nao_encontrado",
                "sem_ptsp15",
            }:
                datas_concluidas.add(data_br)
            else:
                estado["falhas"] += 1

            estado["dias_processados"] = len(datas_concluidas)
            estado["resultados_coletados"] = len(chaves_resultados)
            estado["progresso"] = round(
                (len(datas_concluidas) / PTSP15_TOTAL_DIAS) * 100,
                2
            )

            salvar_estado_ptsp15(estado)

            time.sleep(0.8)
            data_atual += timedelta(days=1)

        estado.update({
            "status": "concluida",
            "data_atual": PTSP15_DATA_FINAL.strftime("%d/%m/%Y"),
            "dias_processados": len(datas_concluidas),
            "resultados_coletados": len(chaves_resultados),
            "progresso": round(
                (len(datas_concluidas) / PTSP15_TOTAL_DIAS) * 100,
                2
            ),
            "mensagem": "Coleta histórica PT-SP 15h concluída.",
        })

        salvar_estado_ptsp15(estado)

    except Exception as e:
        logging.exception("Falha geral na coleta PT-SP 15h.")
        estado.update({
            "status": "erro",
            "mensagem": str(e),
        })
        salvar_estado_ptsp15(estado)

    finally:
        with PTSP15_LOCK:
            PTSP15_THREAD = None


def gerar_excel_ptsp15():
    resultados = carregar_jsonl(ARQUIVO_PTSP15_RESULTADOS)
    auditoria = carregar_jsonl(ARQUIVO_PTSP15_AUDITORIA)

    if not resultados:
        raise ValueError(
            "Nenhum resultado PT-SP 15h foi coletado."
        )

    resultados.sort(
        key=lambda item: datetime.strptime(
            item.get("data", "01/01/1900"),
            "%d/%m/%Y"
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "PT-SP 15H"

    ws.append([
        "Data", "Loteria", "Horário",
        "M1", "M2", "M3", "M4", "M5", "M6", "M7",
        "Origem", "URL",
    ])

    for resultado in resultados:
        ws.append([
            resultado.get("data", ""),
            resultado.get("loteria", "PT-SP"),
            resultado.get("horario", "15"),
            resultado.get("m1", ""),
            resultado.get("m2", ""),
            resultado.get("m3", ""),
            resultado.get("m4", ""),
            resultado.get("m5", ""),
            resultado.get("m6", ""),
            resultado.get("m7", ""),
            resultado.get("origem", ""),
            resultado.get("url", ""),
        ])

    for linha in range(2, ws.max_row + 1):
        ws.cell(linha, 3).number_format = "@"

        for coluna in range(4, 10):
            celula = ws.cell(linha, coluna)
            celula.value = str(celula.value).zfill(4)
            celula.number_format = "@"

        celula_m7 = ws.cell(linha, 10)
        celula_m7.value = str(celula_m7.value).zfill(3)
        celula_m7.number_format = "@"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"

    for col, largura in {
        "A": 13, "B": 12, "C": 10,
        "D": 9, "E": 9, "F": 9, "G": 9, "H": 9, "I": 9, "J": 9,
        "K": 30, "L": 85,
    }.items():
        ws.column_dimensions[col].width = largura

    ws_auditoria = wb.create_sheet("AUDITORIA")
    ws_auditoria.append([
        "Data", "Loteria", "Horário", "Status",
        "Quantidade", "Novos", "URL", "Erro", "Processado em",
    ])

    for registro in auditoria:
        ws_auditoria.append([
            registro.get("data", ""),
            registro.get("loteria", "PT-SP"),
            registro.get("horario", "15"),
            registro.get("status", ""),
            registro.get("quantidade", 0),
            registro.get("novos", 0),
            registro.get("url", ""),
            registro.get("erro", ""),
            registro.get("processado_em", ""),
        ])

    wb.save(ARQUIVO_PTSP15_EXCEL)

    return {
        "resultados": len(resultados),
        "auditoria": len(auditoria),
        "arquivo": str(ARQUIVO_PTSP15_EXCEL),
    }


# ==========================================================
# ROTAS PT-SP 15H
# ==========================================================

@app.route("/ptsp15/teste/<data_teste>")
def teste_ptsp15(data_teste):
    try:
        data_obj = datetime.strptime(
            data_teste,
            "%Y-%m-%d"
        )
    except ValueError:
        return jsonify({
            "ok": False,
            "erro": "Data inválida. Use YYYY-MM-DD.",
        }), 400

    retorno = buscar_ptsp15_data(data_obj)

    return jsonify({
        "ok": True,
        "data_consulta": data_teste,
        **retorno,
    })


@app.route("/ptsp15/debug/<data_teste>")
def debug_ptsp15(data_teste):
    try:
        data_obj = datetime.strptime(
            data_teste,
            "%Y-%m-%d"
        )
    except ValueError:
        return jsonify({
            "ok": False,
            "erro": "Data inválida. Use YYYY-MM-DD.",
        }), 400

    url = montar_url_ptsp15(data_obj)
    resp = requests.get(
        url,
        headers=HEADERS,
        timeout=30,
    )

    estrutura = extrair_dataset_resultadofacil(
        resp.text
    )

    candidatos = []

    for item in estrutura.get("variaveis", []):
        if not isinstance(item, dict):
            continue

        nome = obter_campo_item(item, ["name", "nome"])
        valor = obter_campo_item(item, ["value", "valor"])
        horario = extrair_horario_variavel(nome)
        texto = _normalizar_sem_acentos_ptsp15(nome).upper()

        if (
            horario != "15"
            and "BANDEIR" not in texto
            and "FEDERAL" not in texto
        ):
            continue

        candidatos.append({
            "nome": nome,
            "valor": valor,
            "horario": horario,
            "posicao": extrair_posicao_premio(nome),
            "federal": "FEDERAL" in texto,
            "aceito_ptsp15": nome_variavel_eh_bandeirantes_15(nome),
        })

    return jsonify({
        "ok": True,
        "status_http": resp.status_code,
        "url": url,
        "total_variaveis_dataset": len(
            estrutura.get("variaveis", [])
        ),
        "candidatos": candidatos,
    })


@app.route("/ptsp15/status")
def status_ptsp15():
    estado = carregar_estado_ptsp15()

    return jsonify({
        "ok": True,
        "disco": str(DIRETORIO_DADOS),
        "total_dias_previstos": PTSP15_TOTAL_DIAS,
        "estado": estado,
        "arquivos": {
            "estado": ARQUIVO_PTSP15_ESTADO.exists(),
            "resultados": ARQUIVO_PTSP15_RESULTADOS.exists(),
            "auditoria": ARQUIVO_PTSP15_AUDITORIA.exists(),
            "excel": ARQUIVO_PTSP15_EXCEL.exists(),
        },
    })


@app.route("/ptsp15/iniciar")
def iniciar_ptsp15():
    global PTSP15_THREAD

    with PTSP15_LOCK:
        if (
            PTSP15_THREAD is not None
            and PTSP15_THREAD.is_alive()
        ):
            return jsonify({
                "ok": False,
                "mensagem": "A coleta PT-SP 15h já está em execução.",
                "estado": carregar_estado_ptsp15(),
            }), 409

        PTSP15_THREAD = threading.Thread(
            target=executar_coleta_ptsp15,
            name="coleta_ptsp15",
            daemon=True,
        )
        PTSP15_THREAD.start()

    return jsonify({
        "ok": True,
        "mensagem": "Coleta histórica PT-SP 15h iniciada.",
        "periodo": {
            "inicio": PTSP15_DATA_INICIAL.strftime("%d/%m/%Y"),
            "fim": PTSP15_DATA_FINAL.strftime("%d/%m/%Y"),
        },
        "total_dias": PTSP15_TOTAL_DIAS,
        "status_url": "/ptsp15/status",
    })


@app.route("/ptsp15/gerar-excel")
def rota_gerar_excel_ptsp15():
    estado = carregar_estado_ptsp15()

    if estado.get("status") != "concluida":
        return jsonify({
            "ok": False,
            "mensagem": "A coleta PT-SP 15h ainda não foi concluída.",
            "estado": estado,
        }), 409

    try:
        resumo = gerar_excel_ptsp15()

        return jsonify({
            "ok": True,
            "mensagem": "Planilha PT-SP 15h gerada com sucesso.",
            **resumo,
            "download": "/ptsp15/baixar",
        })

    except Exception as e:
        logging.exception(
            "Erro ao gerar Excel PT-SP 15h."
        )

        return jsonify({
            "ok": False,
            "erro": str(e),
        }), 500


@app.route("/ptsp15/baixar")
def baixar_excel_ptsp15():
    if not ARQUIVO_PTSP15_EXCEL.exists():
        return jsonify({
            "ok": False,
            "mensagem": (
                "A planilha PT-SP 15h ainda não foi gerada. "
                "Acesse /ptsp15/gerar-excel primeiro."
            ),
        }), 404

    return send_file(
        ARQUIVO_PTSP15_EXCEL,
        as_attachment=True,
        download_name=(
            "ptsp_15h_01-01-2025_a_11-08-2026.xlsx"
        ),
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


# ==========================================================
# DEBUG JSON-LD LOTEP 20H - NÃO GRAVA NADA
# ==========================================================

@app.route("/lotep20/debug/<data_teste>")
def debug_lotep20_resultadofacil(
    data_teste
):
    try:
        data_obj = datetime.strptime(
            data_teste,
            "%Y-%m-%d"
        )

    except ValueError:
        return jsonify({
            "ok": False,
            "erro": "Data inválida. Use YYYY-MM-DD.",
        }), 400

    url = montar_url_lotep20(
        data_obj
    )

    try:
        resp = requests.get(
            url,
            headers=HEADERS,
            timeout=30,
        )

    except Exception as e:
        return jsonify({
            "ok": False,
            "erro": str(e),
            "url": url,
        }), 500

    estrutura = extrair_dataset_resultadofacil(
        resp.text
    )

    saida = []

    for item in estrutura.get(
        "variaveis",
        []
    ):
        if not isinstance(
            item,
            dict
        ):
            continue

        nome = obter_campo_item(
            item,
            ["name", "nome"]
        )

        valor = obter_campo_item(
            item,
            ["value", "valor"]
        )

        texto = _normalizar_sem_acentos_lotep20(
            nome
        ).upper()

        # Mostra somente candidatos ligados a LOTEP,
        # Paraíba, PB, PARATODOS ou horário 20.
        if (
            "LOTEP" not in texto
            and "PARATODOS" not in texto
            and "PARA TODOS" not in texto
            and "PARAIBA" not in texto
            and " PB " not in f" {texto} "
            and extrair_horario_variavel(nome) != "20"
        ):
            continue

        saida.append({
            "nome": nome,
            "valor": valor,
            "horario": extrair_horario_variavel(
                nome
            ),
            "posicao": extrair_posicao_premio(
                nome
            ),
            "aceito_lotep20": (
                nome_variavel_eh_paratodos_pb(
                    nome
                )
            ),
        })

    return jsonify({
        "ok": True,
        "status_http": resp.status_code,
        "url": url,
        "total_variaveis_dataset": len(
            estrutura.get(
                "variaveis",
                []
            )
        ),
        "candidatos": saida,
    })



# ==========================================================
# DEBUG HTML LOTEP 20H - NÃO GRAVA NADA
# Procura termos relacionados ao sorteio das 20h no HTML bruto
# e devolve pequenos trechos ao redor de cada ocorrência.
# ==========================================================

@app.route("/lotep20/debug-html/<data_teste>")
def debug_html_lotep20(
    data_teste
):
    try:
        data_obj = datetime.strptime(
            data_teste,
            "%Y-%m-%d"
        )

    except ValueError:
        return jsonify({
            "ok": False,
            "erro": "Data inválida. Use YYYY-MM-DD.",
        }), 400

    url = montar_url_lotep20(
        data_obj
    )

    try:
        resp = requests.get(
            url,
            headers=HEADERS,
            timeout=30,
        )

    except Exception as e:
        return jsonify({
            "ok": False,
            "erro": str(e),
            "url": url,
        }), 500

    html = resp.text or ""

    termos = [
        "PARATODOS",
        "PARA TODOS",
        "20 horas",
        "20h",
        "Paraíba",
        "Paraiba",
        "PB",
    ]

    ocorrencias = []

    html_lower = html.lower()

    for termo in termos:
        termo_lower = termo.lower()

        inicio_busca = 0

        while True:
            pos = html_lower.find(
                termo_lower,
                inicio_busca
            )

            if pos == -1:
                break

            inicio = max(
                0,
                pos - 500
            )

            fim = min(
                len(html),
                pos + len(termo) + 1200
            )

            trecho = html[
                inicio:fim
            ]

            ocorrencias.append({
                "termo": termo,
                "posicao": pos,
                "trecho": trecho,
            })

            inicio_busca = (
                pos + len(termo_lower)
            )

            # Evita resposta gigantesca em páginas com muitas repetições.
            if sum(
                1
                for item in ocorrencias
                if item["termo"] == termo
            ) >= 10:
                break

    # Remove trechos quase duplicados por posição próxima.
    filtradas = []
    posicoes_usadas = []

    for item in sorted(
        ocorrencias,
        key=lambda x: x["posicao"]
    ):
        if any(
            abs(
                item["posicao"] - existente
            ) < 250
            for existente in posicoes_usadas
        ):
            continue

        filtradas.append(
            item
        )

        posicoes_usadas.append(
            item["posicao"]
        )

    # Extrai também o texto visível da página para uma segunda checagem.
    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    texto_visivel = normalizar_texto(
        soup.get_text(
            " ",
            strip=True
        )
    )

    texto_visivel_lower = (
        texto_visivel.lower()
    )

    trechos_texto_visivel = []

    for termo in termos:
        termo_lower = termo.lower()

        inicio_busca = 0

        while True:
            pos = texto_visivel_lower.find(
                termo_lower,
                inicio_busca
            )

            if pos == -1:
                break

            inicio = max(
                0,
                pos - 300
            )

            fim = min(
                len(texto_visivel),
                pos + len(termo) + 900
            )

            trechos_texto_visivel.append({
                "termo": termo,
                "posicao": pos,
                "trecho": texto_visivel[
                    inicio:fim
                ],
            })

            inicio_busca = (
                pos + len(termo_lower)
            )

            if sum(
                1
                for item in trechos_texto_visivel
                if item["termo"] == termo
            ) >= 10:
                break

    return jsonify({
        "ok": True,
        "status_http": resp.status_code,
        "url": url,
        "tamanho_html": len(
            html
        ),
        "total_ocorrencias_html": len(
            filtradas
        ),
        "ocorrencias_html": filtradas,
        "total_ocorrencias_texto": len(
            trechos_texto_visivel
        ),
        "ocorrencias_texto": (
            trechos_texto_visivel
        ),
    })


# ==========================================================
# ROTAS LOTEP 20H
# ==========================================================

@app.route("/lotep20/teste/<data_teste>")
def teste_lotep20(data_teste):
    try:
        data_obj = datetime.strptime(
            data_teste,
            "%Y-%m-%d"
        )
    except ValueError:
        return jsonify({
            "ok": False,
            "erro": "Data inválida. Use YYYY-MM-DD.",
        }), 400

    retorno = buscar_lotep20_data(data_obj)

    return jsonify({
        "ok": True,
        "data_consulta": data_teste,
        **retorno,
    })


@app.route("/lotep20/status")
def status_lotep20():
    estado = carregar_estado_lotep20()

    return jsonify({
        "ok": True,
        "disco": str(DIRETORIO_DADOS),
        "total_dias_previstos": LOTEP20_TOTAL_DIAS,
        "estado": estado,
        "arquivos": {
            "estado": ARQUIVO_LOTEP20_ESTADO.exists(),
            "resultados": ARQUIVO_LOTEP20_RESULTADOS.exists(),
            "auditoria": ARQUIVO_LOTEP20_AUDITORIA.exists(),
            "excel": ARQUIVO_LOTEP20_EXCEL.exists(),
        },
    })


@app.route("/lotep20/iniciar")
def iniciar_lotep20():
    global LOTEP20_THREAD

    with LOTEP20_LOCK:
        if (
            LOTEP20_THREAD is not None
            and LOTEP20_THREAD.is_alive()
        ):
            return jsonify({
                "ok": False,
                "mensagem": (
                    "A coleta PARATODOS PB 09h/20h já está em execução."
                ),
                "estado": carregar_estado_lotep20(),
            }), 409

        LOTEP20_THREAD = threading.Thread(
            target=executar_coleta_lotep20,
            name="coleta_lotep20",
            daemon=True,
        )

        LOTEP20_THREAD.start()

    return jsonify({
        "ok": True,
        "mensagem": (
            "Coleta histórica PARATODOS PB 09h/20h iniciada."
        ),
        "periodo": {
            "inicio": LOTEP20_DATA_INICIAL.strftime("%d/%m/%Y"),
            "fim": LOTEP20_DATA_FINAL.strftime("%d/%m/%Y"),
        },
        "total_dias": LOTEP20_TOTAL_DIAS,
        "status_url": "/lotep20/status",
    })


@app.route("/lotep20/gerar-excel")
def rota_gerar_excel_lotep20():
    estado = carregar_estado_lotep20()

    if estado.get("status") != "concluida":
        return jsonify({
            "ok": False,
            "mensagem": (
                "A coleta PARATODOS PB 09h/20h ainda não foi concluída."
            ),
            "estado": estado,
        }), 409

    try:
        resumo = gerar_excel_lotep20()

        return jsonify({
            "ok": True,
            "mensagem": (
                "Planilha LOTEP 09h/20h gerada com sucesso."
            ),
            **resumo,
            "download": "/lotep20/baixar",
        })

    except Exception as e:
        logging.exception("Erro ao gerar Excel LOTEP 20h.")

        return jsonify({
            "ok": False,
            "erro": str(e),
        }), 500


@app.route("/lotep20/baixar")
def baixar_excel_lotep20():
    if not ARQUIVO_LOTEP20_EXCEL.exists():
        return jsonify({
            "ok": False,
            "mensagem": (
                "A planilha LOTEP 09h/20h ainda não foi gerada. "
                "Acesse /lotep20/gerar-excel primeiro."
            ),
        }), 404

    return send_file(
        ARQUIVO_LOTEP20_EXCEL,
        as_attachment=True,
        download_name=(
            "lotep_09h_20h_01-01-2025_a_12-08-2026.xlsx"
        ),
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


# ==========================================================
# ROTAS DE TESTE
# ==========================================================

@app.route("/diagnostico-lotep-20")
def diagnostico_lotep_20():

    try:

        resultado = (
            diagnosticar_historico_lotep_20_jbcerto()
        )

        return jsonify({
            "ok": True,
            **resultado,
        })

    except Exception as e:

        logging.exception(
            "Erro no diagnóstico histórico "
            "da LOTEP 20h."
        )

        return jsonify({
            "ok": False,
            "erro": str(e),
        }), 500

@app.route("/coleta/status")
def status_coleta():
    estado = carregar_estado_coleta()

    return jsonify({
        "ok": True,
        "disco": str(
            DIRETORIO_DADOS
        ),
        "total_paginas_previstas": (
            TOTAL_PAGINAS_COLETA
        ),
        "estado": estado,
        "arquivos": {
            "estado": ARQUIVO_ESTADO.exists(),
            "resultados": ARQUIVO_RESULTADOS.exists(),
            "auditoria": ARQUIVO_AUDITORIA.exists(),
            "excel": ARQUIVO_EXCEL.exists(),
        },
    })
    
@app.route("/coleta/iniciar")
def iniciar_coleta():
    global COLETA_THREAD

    with COLETA_LOCK:

        if (
            COLETA_THREAD is not None
            and COLETA_THREAD.is_alive()
        ):
            return jsonify({
                "ok": False,
                "mensagem": (
                    "A coleta já está em execução."
                ),
                "estado": carregar_estado_coleta(),
            }), 409

        COLETA_THREAD = threading.Thread(
            target=executar_coleta_historica,
            name="coleta_resultadofacil",
            daemon=True,
        )

        COLETA_THREAD.start()

    return jsonify({
        "ok": True,
        "mensagem": (
            "Coleta histórica iniciada."
        ),
        "periodo": {
            "inicio": "01/01/2025",
            "fim": "12/08/2026",
        },
        "loterias": list(
            LOTERIAS.keys()
        ),
        "total_paginas": (
            TOTAL_PAGINAS_COLETA
        ),
        "status_url": (
            "/coleta/status"
        ),
    })
    
@app.route("/coleta/gerar-excel")
def rota_gerar_excel():

    estado = carregar_estado_coleta()

    if estado.get(
        "status"
    ) != "concluida":

        return jsonify({
            "ok": False,
            "mensagem": (
                "A coleta histórica ainda não "
                "foi concluída."
            ),
            "estado": estado,
        }), 409

    try:
        resumo = gerar_excel_historico()

        return jsonify({
            "ok": True,
            "mensagem": (
                "Planilha gerada com sucesso."
            ),
            **resumo,
            "download": "/coleta/baixar",
        })

    except Exception as e:

        logging.exception(
            "Erro ao gerar Excel histórico."
        )

        return jsonify({
            "ok": False,
            "erro": str(e),
        }), 500
        
@app.route("/coleta/baixar")
def baixar_excel():

    if not ARQUIVO_EXCEL.exists():

        return jsonify({
            "ok": False,
            "mensagem": (
                "A planilha ainda não foi gerada. "
                "Acesse /coleta/gerar-excel primeiro."
            ),
        }), 404

    return send_file(
        ARQUIVO_EXCEL,
        as_attachment=True,
        download_name=(
            "resultadofacil_"
            "01-01-2025_a_12-08-2026.xlsx"
        ),
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
    )

@app.route("/")
def home():
    return jsonify({
        "ok": True,
        "servico": "Coletor Histórico Resultado Fácil",
        "rotas": {
            "/teste/<loteria>/<data>": (
                "Testa uma loteria em uma data"
            ),
            "/teste-periodo": (
                "Testa período curto"
            ),
        },
    })


@app.route(
    "/teste/<loteria>/<data_teste>"
)
def teste_loteria(
    loteria,
    data_teste
):
    loteria = loteria.upper()

    if loteria not in LOTERIAS:
        return jsonify({
            "ok": False,
            "erro": "Loteria inválida",
            "loterias": list(
                LOTERIAS.keys()
            ),
        }), 400

    try:
        data_obj = datetime.strptime(
            data_teste,
            "%Y-%m-%d"
        )

    except ValueError:
        return jsonify({
            "ok": False,
            "erro": (
                "Data inválida. "
                "Use YYYY-MM-DD."
            ),
        }), 400

    retorno = buscar_resultados_data(
        loteria,
        data_obj
    )

    return jsonify(retorno)

@app.route(
    "/debug-html/<loteria>/<data_teste>"
)
def debug_html(
    loteria,
    data_teste
):
    loteria = loteria.upper()

    if loteria not in LOTERIAS:
        return jsonify({
            "ok": False,
            "erro": "Loteria inválida",
        }), 400

    try:
        data_obj = datetime.strptime(
            data_teste,
            "%Y-%m-%d"
        )

    except ValueError:
        return jsonify({
            "ok": False,
            "erro": "Data inválida. Use YYYY-MM-DD.",
        }), 400

    try:
        retorno = diagnosticar_pagina_resultadofacil(
            loteria,
            data_obj
        )

        return jsonify({
            "ok": True,
            **retorno,
        })

    except Exception as e:
        return jsonify({
            "ok": False,
            "erro": str(e),
        }), 500


@app.route("/teste-periodo")
def teste_periodo():
    data_ini = datetime(
        2026,
        8,
        10
    )

    data_fim = datetime(
        2026,
        8,
        12
    )

    saida = []

    data_atual = data_ini

    while data_atual <= data_fim:

        for loteria in LOTERIAS:

            retorno = buscar_resultados_data(
                loteria,
                data_atual
            )

            saida.append({
                "data": data_atual.strftime(
                    "%Y-%m-%d"
                ),
                "loteria": loteria,
                "status": retorno["status"],
                "quantidade": len(
                    retorno["resultados"]
                ),
                "resultados": retorno[
                    "resultados"
                ],
            })

            time.sleep(0.8)

        data_atual += timedelta(
            days=1
        )

    return jsonify({
        "ok": True,
        "total_consultas": len(saida),
        "consultas": saida,
    })


# ==========================================================
# COLETA HISTÓRICA COMPLEMENTAR - PT-SP 09H
# PT-SP 09H -> PT-SP | 09
#
# Criada para recuperar o novo horário das 09h sem depender
# do estado/auditoria da coleta histórica principal.
# O período começa em 01/07/2026 para também auditar datas
# imediatamente anteriores ao primeiro 09h localizado.
# ==========================================================

PTSP09_DATA_INICIAL = datetime(2026, 7, 1)
PTSP09_DATA_FINAL = datetime(2026, 8, 17)
PTSP09_URL_BASE = "https://www.resultadofacil.com.br/"

ARQUIVO_PTSP09_ESTADO = DIRETORIO_DADOS / "ptsp09_estado.json"
ARQUIVO_PTSP09_RESULTADOS = DIRETORIO_DADOS / "ptsp09_resultados.jsonl"
ARQUIVO_PTSP09_AUDITORIA = DIRETORIO_DADOS / "ptsp09_auditoria.jsonl"
ARQUIVO_PTSP09_EXCEL = DIRETORIO_DADOS / "ptsp_09h_2026.xlsx"

PTSP09_LOCK = threading.Lock()
PTSP09_THREAD = None

PTSP09_TOTAL_DIAS = (
    PTSP09_DATA_FINAL.date()
    - PTSP09_DATA_INICIAL.date()
).days + 1


def ptsp09_estado_inicial():
    return {
        "status": "nao_iniciada",
        "data_inicial": PTSP09_DATA_INICIAL.strftime("%d/%m/%Y"),
        "data_final": PTSP09_DATA_FINAL.strftime("%d/%m/%Y"),
        "data_atual": None,
        "dias_processados": 0,
        "resultados_coletados": 0,
        "falhas": 0,
        "progresso": 0.0,
        "mensagem": "A coleta PT-SP 09h ainda não foi iniciada.",
    }


def salvar_estado_ptsp09(estado):
    temporario = ARQUIVO_PTSP09_ESTADO.with_suffix(".tmp")

    with open(temporario, "w", encoding="utf-8") as arquivo:
        json.dump(
            estado,
            arquivo,
            ensure_ascii=False,
            indent=2,
        )

    os.replace(temporario, ARQUIVO_PTSP09_ESTADO)


def carregar_estado_ptsp09():
    if not ARQUIVO_PTSP09_ESTADO.exists():
        estado = ptsp09_estado_inicial()
        salvar_estado_ptsp09(estado)
        return estado

    try:
        with open(
            ARQUIVO_PTSP09_ESTADO,
            "r",
            encoding="utf-8"
        ) as arquivo:
            return json.load(arquivo)

    except Exception:
        logging.exception("Erro ao carregar estado PT-SP 09h.")
        return ptsp09_estado_inicial()


def carregar_datas_ptsp09_concluidas():
    concluidas = set()

    if not ARQUIVO_PTSP09_AUDITORIA.exists():
        return concluidas

    try:
        with open(
            ARQUIVO_PTSP09_AUDITORIA,
            "r",
            encoding="utf-8"
        ) as arquivo:
            for linha in arquivo:
                linha = linha.strip()

                if not linha:
                    continue

                try:
                    registro = json.loads(linha)
                except Exception:
                    continue

                if registro.get("status") not in {
                    "ok",
                    "nao_encontrado",
                    "sem_ptsp09",
                }:
                    continue

                data = registro.get("data", "")

                if data:
                    concluidas.add(data)

    except Exception:
        logging.exception("Erro ao carregar auditoria PT-SP 09h.")

    return concluidas


def carregar_chaves_ptsp09():
    chaves = set()

    if not ARQUIVO_PTSP09_RESULTADOS.exists():
        return chaves

    try:
        with open(
            ARQUIVO_PTSP09_RESULTADOS,
            "r",
            encoding="utf-8"
        ) as arquivo:
            for linha in arquivo:
                linha = linha.strip()

                if not linha:
                    continue

                try:
                    registro = json.loads(linha)
                except Exception:
                    continue

                chaves.add(
                    f"{registro.get('data', '')}|"
                    f"{registro.get('loteria', '')}|"
                    f"{registro.get('horario', '')}"
                )

    except Exception:
        logging.exception("Erro ao carregar resultados PT-SP 09h.")

    return chaves


def montar_url_ptsp09(data_obj):
    data_iso = data_obj.strftime("%Y-%m-%d")

    return (
        f"{PTSP09_URL_BASE}"
        f"resultados-pt-sp-do-dia-"
        f"{data_iso}"
    )


def nome_variavel_eh_ptsp09(nome):
    """
    Aceita apenas o bloco PT-SP/PTSP das 09h.
    Federal e outras bancas são descartadas.
    """
    texto = normalizar_texto(nome).upper()

    if "FEDERAL" in texto:
        return False

    if extrair_horario_variavel(nome) != "09":
        return False

    return (
        "PTSP" in texto
        or "PT SP" in texto
        or "PT-SP" in texto
        or "PTN SP" in texto
    )


def extrair_ptsp09_resultadofacil(html, data_obj, url):
    estrutura = extrair_dataset_resultadofacil(html)
    variaveis = estrutura.get("variaveis", [])

    premios = {}
    nomes_origem = []

    for item in variaveis:
        if not isinstance(item, dict):
            continue

        nome = obter_campo_item(item, ["name", "nome"])
        valor = obter_campo_item(item, ["value", "valor"])

        if not nome or not valor:
            continue

        if not nome_variavel_eh_ptsp09(nome):
            continue

        posicao = extrair_posicao_premio(nome)

        if posicao is None:
            continue

        milhar = extrair_milhar_valor(valor)

        if not milhar:
            continue

        premios[posicao] = milhar
        nomes_origem.append(nome)

    if not all(p in premios for p in range(1, 6)):
        return None

    lista = [
        premios[1],
        premios[2],
        premios[3],
        premios[4],
        premios[5],
    ]

    m6, m7 = calcular_premios_6_7(lista)

    return {
        "data": data_obj.strftime("%d/%m/%Y"),
        "loteria": "PT-SP",
        "horario": "09",
        "m1": premios[1],
        "m2": premios[2],
        "m3": premios[3],
        "m4": premios[4],
        "m5": premios[5],
        "m6": m6,
        "m7": m7,
        "url": url,
        "origem": "RESULTADO_FACIL_PTSP09",
        "variaveis_origem": nomes_origem,
    }


def buscar_ptsp09_data(data_obj):
    url = montar_url_ptsp09(data_obj)

    try:
        resp = requests.get(
            url,
            headers=HEADERS,
            timeout=30,
        )

    except Exception as e:
        return {
            "status": "erro_rede",
            "resultado": None,
            "url": url,
            "erro": str(e),
        }

    if resp.status_code == 404:
        return {
            "status": "nao_encontrado",
            "resultado": None,
            "url": url,
            "erro": "",
        }

    if resp.status_code == 403:
        return {
            "status": "bloqueado_403",
            "resultado": None,
            "url": url,
            "erro": "",
        }

    if resp.status_code != 200:
        return {
            "status": f"http_{resp.status_code}",
            "resultado": None,
            "url": url,
            "erro": "",
        }

    resultado = extrair_ptsp09_resultadofacil(
        resp.text,
        data_obj,
        url,
    )

    if resultado is None:
        return {
            "status": "sem_ptsp09",
            "resultado": None,
            "url": url,
            "erro": "",
        }

    return {
        "status": "ok",
        "resultado": resultado,
        "url": url,
        "erro": "",
    }


def executar_coleta_ptsp09():
    global PTSP09_THREAD

    logging.info(
        "Iniciando coleta histórica complementar PT-SP 09h."
    )

    datas_concluidas = carregar_datas_ptsp09_concluidas()
    chaves_resultados = carregar_chaves_ptsp09()
    estado = carregar_estado_ptsp09()

    estado.update({
        "status": "executando",
        "data_atual": None,
        "dias_processados": len(datas_concluidas),
        "resultados_coletados": len(chaves_resultados),
        "falhas": 0,
        "progresso": round(
            (len(datas_concluidas) / PTSP09_TOTAL_DIAS) * 100,
            2
        ),
        "mensagem": "Coleta histórica PT-SP 09h em andamento.",
    })

    salvar_estado_ptsp09(estado)

    data_atual = PTSP09_DATA_INICIAL

    try:
        while data_atual <= PTSP09_DATA_FINAL:
            data_br = data_atual.strftime("%d/%m/%Y")

            if data_br in datas_concluidas:
                data_atual += timedelta(days=1)
                continue

            estado["data_atual"] = data_br
            estado["mensagem"] = (
                f"Consultando PT-SP 09h em {data_br}."
            )
            salvar_estado_ptsp09(estado)

            retorno = buscar_ptsp09_data(data_atual)
            status = retorno.get("status", "desconhecido")
            resultado = retorno.get("resultado")
            novo = 0

            if status == "ok" and resultado:
                chave = (
                    f"{resultado['data']}|"
                    f"{resultado['loteria']}|"
                    f"{resultado['horario']}"
                )

                if chave not in chaves_resultados:
                    adicionar_jsonl(
                        ARQUIVO_PTSP09_RESULTADOS,
                        resultado
                    )
                    chaves_resultados.add(chave)
                    novo = 1

            adicionar_jsonl(
                ARQUIVO_PTSP09_AUDITORIA,
                {
                    "data": data_br,
                    "loteria": "PT-SP",
                    "horario": "09",
                    "status": status,
                    "quantidade": 1 if resultado else 0,
                    "novos": novo,
                    "url": retorno.get("url", ""),
                    "erro": retorno.get("erro", ""),
                    "processado_em": datetime.now().strftime(
                        "%d/%m/%Y %H:%M:%S"
                    ),
                }
            )

            if status in {
                "ok",
                "nao_encontrado",
                "sem_ptsp09",
            }:
                datas_concluidas.add(data_br)
            else:
                estado["falhas"] += 1

            estado["dias_processados"] = len(datas_concluidas)
            estado["resultados_coletados"] = len(chaves_resultados)
            estado["progresso"] = round(
                (len(datas_concluidas) / PTSP09_TOTAL_DIAS) * 100,
                2
            )

            salvar_estado_ptsp09(estado)

            time.sleep(0.8)
            data_atual += timedelta(days=1)

        estado.update({
            "status": "concluida",
            "data_atual": PTSP09_DATA_FINAL.strftime("%d/%m/%Y"),
            "dias_processados": len(datas_concluidas),
            "resultados_coletados": len(chaves_resultados),
            "progresso": round(
                (len(datas_concluidas) / PTSP09_TOTAL_DIAS) * 100,
                2
            ),
            "mensagem": "Coleta histórica PT-SP 09h concluída.",
        })

        salvar_estado_ptsp09(estado)

    except Exception as e:
        logging.exception("Falha geral na coleta PT-SP 09h.")

        estado.update({
            "status": "erro",
            "mensagem": str(e),
        })

        salvar_estado_ptsp09(estado)

    finally:
        with PTSP09_LOCK:
            PTSP09_THREAD = None


def gerar_excel_ptsp09():
    resultados = carregar_jsonl(ARQUIVO_PTSP09_RESULTADOS)
    auditoria = carregar_jsonl(ARQUIVO_PTSP09_AUDITORIA)

    if not resultados:
        raise ValueError(
            "Nenhum resultado PT-SP 09h foi coletado."
        )

    resultados.sort(
        key=lambda item: datetime.strptime(
            item.get("data", "01/01/1900"),
            "%d/%m/%Y"
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "PT-SP 09H"

    ws.append([
        "Data", "Loteria", "Horário",
        "M1", "M2", "M3", "M4", "M5", "M6", "M7",
        "Origem", "URL",
    ])

    for resultado in resultados:
        ws.append([
            resultado.get("data", ""),
            resultado.get("loteria", "PT-SP"),
            resultado.get("horario", "09"),
            resultado.get("m1", ""),
            resultado.get("m2", ""),
            resultado.get("m3", ""),
            resultado.get("m4", ""),
            resultado.get("m5", ""),
            resultado.get("m6", ""),
            resultado.get("m7", ""),
            resultado.get("origem", ""),
            resultado.get("url", ""),
        ])

    for linha in range(2, ws.max_row + 1):
        ws.cell(linha, 3).number_format = "@"

        for coluna in range(4, 10):
            celula = ws.cell(linha, coluna)
            celula.value = str(celula.value).zfill(4)
            celula.number_format = "@"

        celula_m7 = ws.cell(linha, 10)
        celula_m7.value = str(celula_m7.value).zfill(3)
        celula_m7.number_format = "@"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"

    for col, largura in {
        "A": 13, "B": 12, "C": 10,
        "D": 9, "E": 9, "F": 9, "G": 9, "H": 9, "I": 9, "J": 9,
        "K": 28, "L": 85,
    }.items():
        ws.column_dimensions[col].width = largura

    ws_auditoria = wb.create_sheet("AUDITORIA")

    ws_auditoria.append([
        "Data", "Loteria", "Horário", "Status",
        "Quantidade", "Novos", "URL", "Erro", "Processado em",
    ])

    for registro in auditoria:
        ws_auditoria.append([
            registro.get("data", ""),
            registro.get("loteria", "PT-SP"),
            registro.get("horario", "09"),
            registro.get("status", ""),
            registro.get("quantidade", 0),
            registro.get("novos", 0),
            registro.get("url", ""),
            registro.get("erro", ""),
            registro.get("processado_em", ""),
        ])

    ws_auditoria.freeze_panes = "A2"
    ws_auditoria.auto_filter.ref = (
        f"A1:I{ws_auditoria.max_row}"
    )

    wb.save(ARQUIVO_PTSP09_EXCEL)

    return {
        "resultados": len(resultados),
        "auditoria": len(auditoria),
        "arquivo": str(ARQUIVO_PTSP09_EXCEL),
    }


# ==========================================================
# ROTAS PT-SP 09H
# ==========================================================

@app.route("/ptsp09/teste/<data_teste>")
def teste_ptsp09(data_teste):
    try:
        data_obj = datetime.strptime(
            data_teste,
            "%Y-%m-%d"
        )
    except ValueError:
        return jsonify({
            "ok": False,
            "erro": "Data inválida. Use YYYY-MM-DD.",
        }), 400

    retorno = buscar_ptsp09_data(data_obj)

    return jsonify({
        "ok": True,
        "data_consulta": data_teste,
        **retorno,
    })


@app.route("/ptsp09/debug/<data_teste>")
def debug_ptsp09(data_teste):
    try:
        data_obj = datetime.strptime(
            data_teste,
            "%Y-%m-%d"
        )
    except ValueError:
        return jsonify({
            "ok": False,
            "erro": "Data inválida. Use YYYY-MM-DD.",
        }), 400

    url = montar_url_ptsp09(data_obj)

    resp = requests.get(
        url,
        headers=HEADERS,
        timeout=30,
    )

    estrutura = extrair_dataset_resultadofacil(
        resp.text
    )

    candidatos = []

    for item in estrutura.get("variaveis", []):
        if not isinstance(item, dict):
            continue

        nome = obter_campo_item(item, ["name", "nome"])
        valor = obter_campo_item(item, ["value", "valor"])
        horario = extrair_horario_variavel(nome)
        texto = normalizar_texto(nome).upper()

        if (
            horario != "09"
            and "PTSP" not in texto
            and "PT SP" not in texto
        ):
            continue

        candidatos.append({
            "nome": nome,
            "valor": valor,
            "horario": horario,
            "posicao": extrair_posicao_premio(nome),
            "federal": "FEDERAL" in texto,
            "aceito_ptsp09": nome_variavel_eh_ptsp09(nome),
        })

    return jsonify({
        "ok": True,
        "status_http": resp.status_code,
        "url": url,
        "total_variaveis_dataset": len(
            estrutura.get("variaveis", [])
        ),
        "candidatos": candidatos,
    })


@app.route("/ptsp09/status")
def status_ptsp09():
    estado = carregar_estado_ptsp09()

    return jsonify({
        "ok": True,
        "disco": str(DIRETORIO_DADOS),
        "total_dias_previstos": PTSP09_TOTAL_DIAS,
        "estado": estado,
        "arquivos": {
            "estado": ARQUIVO_PTSP09_ESTADO.exists(),
            "resultados": ARQUIVO_PTSP09_RESULTADOS.exists(),
            "auditoria": ARQUIVO_PTSP09_AUDITORIA.exists(),
            "excel": ARQUIVO_PTSP09_EXCEL.exists(),
        },
    })


@app.route("/ptsp09/iniciar")
def iniciar_ptsp09():
    global PTSP09_THREAD

    with PTSP09_LOCK:
        if (
            PTSP09_THREAD is not None
            and PTSP09_THREAD.is_alive()
        ):
            return jsonify({
                "ok": False,
                "mensagem": "A coleta PT-SP 09h já está em execução.",
                "estado": carregar_estado_ptsp09(),
            }), 409

        PTSP09_THREAD = threading.Thread(
            target=executar_coleta_ptsp09,
            name="coleta_ptsp09",
            daemon=True,
        )

        PTSP09_THREAD.start()

    return jsonify({
        "ok": True,
        "mensagem": "Coleta histórica PT-SP 09h iniciada.",
        "periodo": {
            "inicio": PTSP09_DATA_INICIAL.strftime("%d/%m/%Y"),
            "fim": PTSP09_DATA_FINAL.strftime("%d/%m/%Y"),
        },
        "total_dias": PTSP09_TOTAL_DIAS,
        "status_url": "/ptsp09/status",
    })


@app.route("/ptsp09/gerar-excel")
def rota_gerar_excel_ptsp09():
    estado = carregar_estado_ptsp09()

    if estado.get("status") != "concluida":
        return jsonify({
            "ok": False,
            "mensagem": "A coleta PT-SP 09h ainda não foi concluída.",
            "estado": estado,
        }), 409

    try:
        resumo = gerar_excel_ptsp09()

        return jsonify({
            "ok": True,
            "mensagem": "Planilha PT-SP 09h gerada com sucesso.",
            **resumo,
            "download": "/ptsp09/baixar",
        })

    except Exception as e:
        logging.exception(
            "Erro ao gerar Excel PT-SP 09h."
        )

        return jsonify({
            "ok": False,
            "erro": str(e),
        }), 500


@app.route("/ptsp09/baixar")
def baixar_excel_ptsp09():
    if not ARQUIVO_PTSP09_EXCEL.exists():
        return jsonify({
            "ok": False,
            "mensagem": (
                "A planilha PT-SP 09h ainda não foi gerada. "
                "Acesse /ptsp09/gerar-excel primeiro."
            ),
        }), 404

    return send_file(
        ARQUIVO_PTSP09_EXCEL,
        as_attachment=True,
        download_name=(
            "ptsp_09h_01-07-2026_a_17-08-2026.xlsx"
        ),
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


if __name__ == "__main__":
    porta = int(
        os.environ.get(
            "PORT",
            5000
        )
    )

    app.run(
        host="0.0.0.0",
        port=porta
    )
